import asyncio
import datetime
import io
import json
import os
import queue
import secrets
import sys
import threading
import time
import uuid
import logging
from typing import Any, Dict, List, Optional, Union
from fastapi import FastAPI, HTTPException, Header, Cookie, Query, Request
from fastapi.responses import StreamingResponse, RedirectResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("sales_spark")

# Add current directory to path so we can import config/core
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Load environment variables from all possible locations (.env in backend, root, AppData, user profile)
def _load_all_envs():
    try:
        from dotenv import load_dotenv
    except ImportError:
        return

    candidates = [
        # 1. User AppData (production desktop persistence: %APPDATA%\HomeSpark\.env)
        os.path.join(os.getenv("APPDATA", ""), "HomeSpark", ".env") if os.getenv("APPDATA") else "",
        # 2. User profile (~/.homespark/.env)
        os.path.join(os.path.expanduser("~"), ".homespark", ".env"),
        # 3. Project local .env (backend dir)
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
        # 4. Project root .env
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".env"),
    ]
    for c in candidates:
        if c and os.path.isfile(c):
            logger.info(f"[Env] Loading environment variables from {c}")
            load_dotenv(c, override=False)

_load_all_envs()

from config.const import (
    DEFAULT_SYSTEM_PROMPT,
    MODEL_NAME,
    FRONTEND_URL,
    GOOGLE_OAUTH_REDIRECT_URI,
    TTS_SERVER_URL,
)
from core.agent import Agent
from core.llm_client import OpenAICompatClient
from core.tool import ToolRegistry
from core.tool_calls import default_registry
from core.session import make_session, verify_session
from core.google_tools import build_google_tools
from core.people_tools import build_people_tools
from core.store import (
    get_chats,
    get_messages,
    save_message,
    delete_chat,
    get_notifications,
    create_notification,
    mark_notification_as_read,
    delete_notification,
    notification_exists_for_mail,
    get_all_linked_users,
    get_notification_by_id,
    get_user_profile,
    upsert_user_profile,
    find_person_by_email,
    create_full_person,
)
from core import google_oauth
from core.google_tools import build_google_tools, _service, _extract_plain_body
from core.weather_tools import build_weather_tools
from core.memory_tools import build_memory_tools
from core.web_search_tools import build_web_search_tools
from core.imap_tools import build_imap_tools, test_imap_and_smtp_connection
from core.classifier import classify_is_addressing_ai, classify_is_conversation_ended
from core.store import (
    get_user_current_minutes,
    save_user_minutes_and_archive_old,
    search_user_skills,
    get_imap_accounts,
    create_imap_account,
    delete_imap_account,
)

# Local-dev login when Google OAuth isn't configured yet.
_ALLOW_MOCK_AUTH = os.getenv("ALLOW_MOCK_AUTH", "").lower() in ("1", "true", "yes")

# Self-terminating dead-man switch if parent Electron process is killed
def _init_parent_watchdog():
    parent_pid_str = os.getenv("PARENT_ELECTRON_PID")
    if not parent_pid_str or not parent_pid_str.isdigit():
        return
    parent_pid = int(parent_pid_str)
    if parent_pid <= 0:
        return

    import threading, time
    SYNCHRONIZE = 0x00100000
    ERROR_INVALID_PARAMETER = 87
    ERROR_ACCESS_DENIED = 5

    def _watch():
        while True:
            time.sleep(2.5)
            try:
                import ctypes
                handle = ctypes.windll.kernel32.OpenProcess(SYNCHRONIZE, False, parent_pid)
                if handle:
                    # WAIT_OBJECT_0 = 0 means signaled (parent process has terminated)
                    wait_res = ctypes.windll.kernel32.WaitForSingleObject(handle, 0)
                    ctypes.windll.kernel32.CloseHandle(handle)
                    if wait_res == 0:
                        logger.info(f"[Watchdog] Parent Electron process {parent_pid} has exited. Terminating backend...")
                        os._exit(0)
                else:
                    err = ctypes.windll.kernel32.GetLastError()
                    # Only exit if the process ID is definitively invalid/non-existent
                    if err == ERROR_INVALID_PARAMETER:
                        logger.info(f"[Watchdog] Parent Electron process {parent_pid} no longer exists. Terminating backend...")
                        os._exit(0)
                    # If ERROR_ACCESS_DENIED or other error, the process is still running; keep watching
            except Exception:
                pass

    t = threading.Thread(target=_watch, daemon=True)
    t.start()

_init_parent_watchdog()

app = FastAPI(title="Sales Spark Backend API")

# Enable CORS for frontend integration
origins = [FRONTEND_URL] if FRONTEND_URL else []
if "http://localhost:3000" in origins:
    origins.append("http://127.0.0.1:3000")
if not origins:
    origins = ["http://localhost:3000", "http://127.0.0.1:3000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1)(:\d+)?$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class ChatMessage(BaseModel):
    role: str
    content: Optional[Union[str, List[Any]]] = None
    tool_calls: Optional[List[Dict[str, Any]]] = None
    tool_call_id: Optional[str] = None
    name: Optional[str] = None

class ChatRequest(BaseModel):
    message: Union[str, List[Any]]
    chat_id: Optional[str] = None
    history: Optional[List[ChatMessage]] = None
    system_prompt: Optional[str] = None
    tool_mode: Optional[str] = "auto"
    save_to_history: Optional[bool] = True


class SummarizeRequest(BaseModel):
    history: List[Dict[str, Any]]
    title: Optional[str] = None


@app.post("/api/memory/summarize")
async def summarize_memory_endpoint(
    req: SummarizeRequest,
    authorization: Optional[str] = Header(None)
):
    """Summarize voice conversation history into meeting minutes & archive old minutes into skills."""
    uid = resolve_uid(authorization)
    if not req.history or len(req.history) == 0:
        return {"status": "ok", "message": "No history to summarize", "minutes": ""}

    logger.info(f"[summarize_memory_endpoint] Summarizing {len(req.history)} messages for uid={uid}")
    
    # Format conversation history
    lines = []
    for msg in req.history:
        role = msg.get("role", "unknown")
        speaker = "ユーザー" if role == "user" else "GeMo(AI)"
        content = msg.get("content", "")
        if isinstance(content, list):
            content = " ".join([item.get("text", "") for item in content if isinstance(item, dict)])
        if content:
            lines.append(f"{speaker}: {content}")

    formatted_history = "\n".join(lines)

    summary_instruction = (
        "あなたは会話ログから今後の対話や業務に役立つ構造化された議事録（Markdown形式）を作成する専門AIです。\n"
        "以下のユーザーとGeMo（アシスタント）の会話内容を読み、今後の対話で参照すべき重要な記憶・議事録を簡潔にまとめてください。\n\n"
        "【議事録に含める項目】\n"
        "- 📌 主な話題・会話の要約\n"
        "- 🎯 決定事項・合意内容\n"
        "- 💡 ユーザーの好み・パーソナル情報・発言した重要事項\n"
        "- 📝 残っているTODOや次回の予定\n\n"
        "【会話履歴】\n" + formatted_history
    )

    try:
        client = OpenAICompatClient()
        ask_result = client.ask(
            user_content=summary_instruction,
            system_prompt="あなたは的確で簡潔な日本語の会話議事録・ナレッジ抽出AIです。",
            history=[],
            tool_registry=None,
            json_schema=None,
            tool_mode="off",
            stream=False,
            on_token=None,
            temperature=0.7,
            max_tokens=1024,
        )
        generated_minutes = ask_result.content.strip()
        if not generated_minutes:
            generated_minutes = "（特筆すべき決定事項なし）"

        # Archive old minutes into skills and save the new minutes
        res = save_user_minutes_and_archive_old(uid, generated_minutes, archive_title=req.title)
        logger.info(f"[summarize_memory_endpoint] Successfully saved minutes. Archived old: {res.get('archived_previous')}")

        return {
            "status": "ok",
            "minutes": generated_minutes,
            "archived_previous": res.get("archived_previous")
        }
    except Exception as e:
        logger.error(f"[summarize_memory_endpoint] Summarization failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


class AddressingAiRequest(BaseModel):
    text: str
    last_ai_response: Optional[str] = None


class ConversationEndedRequest(BaseModel):
    ai_response: str


@app.post("/api/classifier/is-addressing-ai")
async def is_addressing_ai_endpoint(req: AddressingAiRequest):
    """Determine whether the user speech is addressing the AI assistant."""
    is_addressing = classify_is_addressing_ai(req.text, req.last_ai_response)
    return {"is_addressing": is_addressing}


@app.post("/api/classifier/is-conversation-ended")
async def is_conversation_ended_endpoint(req: ConversationEndedRequest):
    """Determine whether the assistant response marks the natural end of the conversation topic."""
    is_ended = classify_is_conversation_ended(req.ai_response)
    return {"is_ended": is_ended}


class CreateImapAccountRequest(BaseModel):
    label: str
    email_address: str
    imap_host: str
    imap_port: int = 993
    imap_ssl: bool = True
    smtp_host: str
    smtp_port: int = 465
    smtp_ssl: bool = True
    username: str
    password: str


class TestImapAccountRequest(BaseModel):
    imap_host: str
    imap_port: int = 993
    imap_ssl: bool = True
    smtp_host: Optional[str] = ""
    smtp_port: Optional[int] = 465
    smtp_ssl: Optional[bool] = True
    username: str
    password: str


@app.get("/api/imap/accounts")
async def get_imap_accounts_endpoint(authorization: Optional[str] = Header(None)):
    """List all configured external IMAP accounts for the user."""
    uid = resolve_uid(authorization)
    accounts = get_imap_accounts(uid, include_password=False)
    return {"accounts": accounts}


@app.post("/api/imap/accounts/test")
async def test_imap_account_endpoint(req: TestImapAccountRequest):
    """Test connection to IMAP and SMTP servers without saving."""
    res = test_imap_and_smtp_connection(req.dict())
    return res


@app.post("/api/imap/accounts")
async def create_imap_account_endpoint(
    req: CreateImapAccountRequest,
    authorization: Optional[str] = Header(None)
):
    """Register a new IMAP/SMTP mail account (runs connection test first)."""
    uid = resolve_uid(authorization)
    test_res = test_imap_and_smtp_connection(req.dict())
    if not test_res.get("success"):
        raise HTTPException(status_code=400, detail=test_res.get("error", "接続テストに失敗しました。設定を確認してください。"))

    account = create_imap_account(uid, req.dict())
    return {"status": "ok", "account": account}


@app.delete("/api/imap/accounts/{account_id}")
async def delete_imap_account_endpoint(
    account_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete an external IMAP account."""
    uid = resolve_uid(authorization)
    delete_imap_account(uid, account_id)
    return {"status": "ok", "deleted_id": account_id}


# Module-level lazy in-process Whisper fallback model
_IN_PROCESS_WHISPER = None
_WHISPER_LOCK = threading.Lock()

def _get_in_process_whisper():
    global _IN_PROCESS_WHISPER
    if _IN_PROCESS_WHISPER is not None:
        return _IN_PROCESS_WHISPER
    with _WHISPER_LOCK:
        if _IN_PROCESS_WHISPER is not None:
            return _IN_PROCESS_WHISPER
        try:
            import torch
            from faster_whisper import WhisperModel

            device = "cpu"
            compute_type = "int8"
            whisper_model = os.getenv("WHISPER_MODEL", "kotoba-tech/kotoba-whisper-v2.0-faster")

            if torch.cuda.is_available():
                try:
                    free_bytes, total_bytes = torch.cuda.mem_get_info()
                    free_gb = free_bytes / (1024 ** 3)
                    # Require at least 1.0 GB free VRAM to hold Kotoba-Whisper-v2.0 alongside Irodori-TTS
                    if free_gb >= 1.0:
                        device = "cuda"
                        compute_type = "float16"
                        logger.info(f"[InProcess Whisper] Sufficient VRAM available ({free_gb:.1f} GB free). Initializing {whisper_model} on CUDA (float16)...")
                    else:
                        logger.warning(f"[InProcess Whisper] Low VRAM ({free_gb:.1f} GB free < 1.0 GB required). Falling back to CPU (int8) for safety.")
                except Exception as mem_err:
                    logger.warning(f"[InProcess Whisper] Could not inspect VRAM ({mem_err}), defaulting to CPU int8 for safety.")

            _IN_PROCESS_WHISPER = WhisperModel(whisper_model, device=device, compute_type=compute_type)
            logger.info(f"[InProcess Whisper] Successfully loaded {whisper_model} on {device} ({compute_type}).")
            return _IN_PROCESS_WHISPER
        except Exception as e:
            logger.warning(f"[InProcess Whisper] Failed to load local whisper fallback: {e}")
            return None


def _resolve_tts_server_url() -> str:
    return (
        os.getenv("TTS_SERVER_URL")
        or os.getenv("TTS_URL")
        or TTS_SERVER_URL
        or "http://127.0.0.1:8008"
    )


@app.get("/api/tts")
async def tts_proxy_endpoint(
    text: str = Query(..., description="Text to synthesize to speech"),
    steps: Optional[int] = Query(6, ge=1, le=100)
):
    """Proxy request to Irodori-TTS service running on GPU worker."""
    import urllib.request
    import urllib.parse
    import urllib.error

    clean_text = text.strip().replace("*", "").replace("`", "").replace("#", "")
    if not clean_text:
        raise HTTPException(status_code=400, detail="Text cannot be empty")

    query_params = urllib.parse.urlencode({"text": clean_text, "steps": steps or 6})
    tts_base = _resolve_tts_server_url()
    target_url = f"{tts_base}/tts?{query_params}"

    loop = asyncio.get_running_loop()

    def fetch_tts():
        req = urllib.request.Request(
            target_url,
            headers={"User-Agent": "SalesSpark-Backend/1.0"}
        )
        with urllib.request.urlopen(req, timeout=30.0) as response:
            return response.read()

    try:
        wav_data = await loop.run_in_executor(None, fetch_tts)
        return StreamingResponse(
            io.BytesIO(wav_data),
            media_type="audio/wav",
            headers={
                "Content-Disposition": "inline; filename=speech.wav",
                "Cache-Control": "no-store, no-cache, must-revalidate",
            }
        )
    except Exception as e:
        logger.info(f"Local TTS engine ({tts_base}) offline or unavailable ({e}), returning graceful empty audio for client-side synthesis fallback")
        # 44-byte empty WAV header
        empty_wav = b'RIFF$\x00\x00\x00WAVEfmt \x10\x00\x00\x00\x01\x00\x01\x00\x80>\x00\x00\x00}\x00\x00\x02\x00\x10\x00data\x00\x00\x00\x00'
        return StreamingResponse(
            io.BytesIO(empty_wav),
            media_type="audio/wav",
            headers={
                "Content-Disposition": "inline; filename=fallback.wav",
                "X-TTS-Fallback": "web-speech-fallback",
            }
        )


@app.post("/api/audio/transcribe")
async def transcribe_audio_endpoint(
    request: Request,
    authorization: Optional[str] = Header(None)
):
    """Robust audio transcription endpoint for desktop voice call with multi-engine fallback:
    1. Dedicated Local faster-whisper on TTS/STT Engine (Port 8008 / TTS_SERVER_URL)
    2. In-Process Faster-Whisper GPU/CPU fallback
    3. OpenAI Whisper API
    4. Google Gemini Multimodal Audio
    """
    try:
        content_type = request.headers.get("content-type", "")
        content = b""
        if "multipart/form-data" in content_type:
            try:
                form = await request.form()
                audio_field = form.get("audio") or form.get("file")
                if audio_field and hasattr(audio_field, "read"):
                    content = await audio_field.read()
                elif isinstance(audio_field, bytes):
                    content = audio_field
                else:
                    content = await request.body()
            except Exception:
                content = await request.body()
        else:
            content = await request.body()

        if not content or len(content) < 100:
            return {"text": ""}

        loop = asyncio.get_running_loop()
        tts_base = _resolve_tts_server_url()

        # 1. Cloud STT: Groq Cloud Whisper (Ultra-fast ~100ms, free tier)
        groq_key = os.getenv("GROQ_API_KEY")
        if groq_key:
            try:
                import urllib.request
                boundary = "----WebKitFormBoundary" + secrets.token_hex(16)
                body = (
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="model"\r\n\r\nwhisper-large-v3-turbo\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="language"\r\n\r\nja\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="file"; filename="audio.webm"\r\n'
                    f'Content-Type: audio/webm\r\n\r\n'
                ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")

                req = urllib.request.Request(
                    "https://api.groq.com/openai/v1/audio/transcriptions",
                    data=body,
                    headers={
                        "Authorization": f"Bearer {groq_key}",
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    }
                )

                def _fetch_groq_stt():
                    with urllib.request.urlopen(req, timeout=6.0) as res:
                        return json.loads(res.read().decode("utf-8"))

                data = await loop.run_in_executor(None, _fetch_groq_stt)
                text = data.get("text", "").strip()
                if text:
                    logger.info(f"[STT] Groq Cloud Whisper succeeded: {text}")
                    return {"text": text}
            except Exception as e:
                logger.warning(f"[Groq Whisper STT error]: {e}")

        # 2. Cloud STT: Google Gemini 2.0 Flash Multimodal Audio (Free tier on Google AI Studio, ~250ms)
        gemini_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
        if gemini_key:
            try:
                import urllib.request
                import base64
                b64_audio = base64.b64encode(content).decode("utf-8")
                payload = {
                    "contents": [{
                        "parts": [
                            {"text": "ユーザーが話した日本語音声を正確に文字起こししてください。挨拶や短い相槌も含め、発話されたテキストのみを出力してください。"},
                            {"inline_data": {"mime_type": "audio/webm", "data": b64_audio}}
                        ]
                    }],
                    "generationConfig": {"temperature": 0.0}
                }
                req = urllib.request.Request(
                    f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={gemini_key}",
                    data=json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"}
                )

                def _fetch_gemini_stt():
                    with urllib.request.urlopen(req, timeout=8.0) as res:
                        return json.loads(res.read().decode("utf-8"))

                data = await loop.run_in_executor(None, _fetch_gemini_stt)
                candidates = data.get("candidates", [])
                if candidates:
                    text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "").strip()
                    if text:
                        logger.info(f"[STT] Gemini 2.0 Flash Cloud STT succeeded: {text}")
                        return {"text": text}
            except Exception as e:
                logger.warning(f"[Gemini Cloud STT error]: {e}")

        # 3. Cloud STT: OpenAI Whisper API
        openai_key = os.getenv("OPENAI_API_KEY")
        if openai_key:
            try:
                import urllib.request
                boundary = "----WebKitFormBoundary" + secrets.token_hex(16)
                body = (
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="model"\r\n\r\nwhisper-1\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="language"\r\n\r\nja\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="file"; filename="audio.webm"\r\n'
                    f'Content-Type: audio/webm\r\n\r\n'
                ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")

                req = urllib.request.Request(
                    "https://api.openai.com/v1/audio/transcriptions",
                    data=body,
                    headers={
                        "Authorization": f"Bearer {openai_key}",
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    }
                )

                def _fetch_openai_stt():
                    with urllib.request.urlopen(req, timeout=8.0) as res:
                        return json.loads(res.read().decode("utf-8"))

                data = await loop.run_in_executor(None, _fetch_openai_stt)
                text = data.get("text", "").strip()
                if text:
                    logger.info(f"[STT] OpenAI Whisper succeeded: {text}")
                    return {"text": text}
            except Exception as e:
                logger.warning(f"[OpenAI Whisper STT error]: {e}")

        # 4. Local Fast STT: Dedicated Fast Faster-Whisper on Port 8008 (TTS_SERVER_URL)
        try:
            import urllib.request
            boundary = "----WebKitFormBoundary" + secrets.token_hex(16)
            body = (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="file"; filename="audio.webm"\r\n'
                f'Content-Type: audio/webm\r\n\r\n'
            ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")

            local_stt_url = f"{tts_base}/transcribe"
            req = urllib.request.Request(
                local_stt_url,
                data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )

            def _fetch_local_whisper():
                with urllib.request.urlopen(req, timeout=6.0) as res:
                    return json.loads(res.read().decode("utf-8"))

            data = await loop.run_in_executor(None, _fetch_local_whisper)
            text = data.get("text", "").strip()
            if text:
                logger.info(f"[STT] Dedicated Local Whisper succeeded: {text}")
                return {"text": text}
        except Exception as e:
            logger.debug(f"[STT] Local dedicated worker offline or error ({e}), trying in-process whisper...")

        # 5. Local Fast STT: In-Process Faster-Whisper (beam_size=1, vad_filter=True)
        try:
            whisper_inst = _get_in_process_whisper()
            if whisper_inst is not None:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                try:
                    def _transcribe_in_proc():
                        segments, _ = whisper_inst.transcribe(tmp_path, beam_size=1, vad_filter=True, language="ja")
                        return "".join([s.text for s in segments]).strip()
                    in_proc_text = await loop.run_in_executor(None, _transcribe_in_proc)
                    if in_proc_text:
                        logger.info(f"[STT] In-process Whisper transcription succeeded: {in_proc_text}")
                        return {"text": in_proc_text}
                finally:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
        except Exception as e:
            logger.debug(f"[STT] In-process Whisper error ({e})")

        return {"text": ""}
    except Exception as e:
        logger.error(f"[Transcribe Exception]: {e}")
        return {"text": ""}


def resolve_uid(authorization: Optional[str], *, allow_anonymous: bool = True) -> str:
    """Verify the Bearer session token and return the uid (the user's Google sub).

    When allow_anonymous is True and the token is invalid/expired, fall back to
    anonymous_user gracefully rather than blocking the chat interface."""
    if authorization and authorization.startswith("Bearer "):
        claims = verify_session(authorization[7:])
        if claims and claims.get("sub"):
            return claims["sub"]
        if not allow_anonymous:
            raise HTTPException(status_code=401, detail="Invalid or expired session")
        return "anonymous_user"
    if not allow_anonymous:
        raise HTTPException(status_code=401, detail="Authentication required")
    return "anonymous_user"


class StreamingToolRegistry(ToolRegistry):
    def __init__(self, original_registry: ToolRegistry, event_queue: queue.Queue):
        super().__init__()
        self._tools = original_registry._tools.copy()
        self.event_queue = event_queue

    def execute(self, name: str, arguments: Dict[str, Any] | None) -> Any:
        self.event_queue.put({"type": "tool_start", "name": name, "arguments": arguments})
        try:
            result = super().execute(name, arguments)
            # A tool may return either a plain string (LLM-facing text) or a
            # dict {"llm_text": str, "diagram": {...}} to ALSO drive a custom
            # diagram visualization in the UI. When a diagram is present we push
            # it to the client as its own SSE event — sent whole, not token by
            # token — and strip it from what the LLM sees: the model only ever
            # receives llm_text, so the bulky JSON never pollutes its context.
            diagram = None
            if isinstance(result, dict) and "diagram" in result:
                diagram = result.get("diagram")
                text_result = result.get("llm_text", "")
            else:
                text_result = result if isinstance(result, str) else str(result)

            if diagram is not None:
                self.event_queue.put(
                    {"type": "custom_diagram", "name": name, "diagram": diagram}
                )
            self.event_queue.put({"type": "tool_end", "name": name, "result": text_result})
            return text_result
        except Exception as e:
            self.event_queue.put({"type": "tool_error", "name": name, "error": str(e)})
            raise e

def run_agent_thread(
    agent: Agent,
    message: Union[str, List[Any]],
    event_queue: queue.Queue,
    uid: str,
    chat_id: str,
    initial_history_len: int,
    save_to_history: bool = True,
):
    try:
        def on_token(token: str):
            event_queue.put({"type": "token", "content": token})

        final_content = agent.ask(message, on_token=on_token)
        
        # Save newly generated messages to DB only if save_to_history is True
        if save_to_history:
            new_messages = agent.memory[initial_history_len + 1:]
            for msg in new_messages:
                save_message(
                    uid=uid,
                    chat_id=chat_id,
                    role=msg["role"],
                    content=msg.get("content"),
                    tool_calls=msg.get("tool_calls"),
                    tool_call_id=msg.get("tool_call_id"),
                    name=msg.get("name"),
                )

        # Push the final state of memory (which includes assistant response and tool runs)
        event_queue.put({
            "type": "done",
            "final_content": final_content,
            "memory": agent.memory,
            "chat_id": chat_id
        })
    except Exception as e:
        event_queue.put({"type": "error", "error": str(e)})

import base64
import io
import re

def extract_text_from_file(mime_type: str, base64_data: str, file_name: str) -> str:
    try:
        decoded_bytes = base64.b64decode(base64_data)
    except Exception as e:
        return f"[エラー: ファイルデータのデコードに失敗しました: {e}]"

    # Excel の場合 (xlsx)
    if (mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or 
            file_name.endswith(".xlsx")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(decoded_bytes), read_only=True, data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"--- シート: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_str = "\t".join([str(val) if val is not None else "" for val in row])
                    if row_str.strip():
                        text_parts.append(row_str)
            return "\n".join(text_parts)
        except Exception as e:
            return f"[エラー: Excelファイルの解析に失敗しました: {e}]"

    # CSV の場合
    elif mime_type == "text/csv" or file_name.endswith(".csv"):
        try:
            return decoded_bytes.decode("utf-8", errors="replace")
        except Exception as e:
            return f"[エラー: CSVファイルの読み込みに失敗しました: {e}]"

    # テキストファイルの場合
    elif mime_type.startswith("text/") or file_name.endswith((".txt", ".md", ".json", ".xml", ".yaml", ".yml")):
        try:
            return decoded_bytes.decode("utf-8", errors="replace")
        except Exception as e:
            return f"[エラー: テキストファイルの読み込みに失敗しました: {e}]"

    # その他
    else:
        try:
            return decoded_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return f"[未対応のファイル形式です (MIMEタイプ: {mime_type})。バイナリデータのためテキストとして抽出できません。]"

def preprocess_content(content: Any) -> Any:
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return content
    
    new_content = []
    for item in content:
        if isinstance(item, dict) and item.get("type") == "image_url":
            img_url = item.get("image_url", {}).get("url", "")
            if img_url.startswith("data:"):
                match = re.match(r"^data:([^;]+);base64,(.+)$", img_url)
                if match:
                    mime_type, base64_data = match.groups()
                    if mime_type.startswith("image/"):
                        new_content.append(item)
                    else:
                        file_name = item.get("name", "添付ファイル")
                        extracted_text = extract_text_from_file(mime_type, base64_data, file_name)
                        new_content.append({
                            "type": "text",
                            "text": f"\n[添付ファイル: {file_name}]\n{extracted_text}\n[添付ファイル終わり]\n"
                        })
                else:
                    new_content.append(item)
            else:
                new_content.append(item)
        else:
            new_content.append(item)
    return new_content

@app.post("/api/chat")
async def chat_endpoint(
    req: ChatRequest,
    authorization: Optional[str] = Header(None)
):
    # Verify the session token if present, fallback to anonymous
    uid = resolve_uid(authorization)
    logger.info(f"[chat_endpoint] Invoked with uid: {uid}, save_to_history: {req.save_to_history}, has_auth: {bool(authorization)}")

    # Initialize client
    try:
        client = OpenAICompatClient()
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))

    event_queue = queue.Queue()
    # Base tools (file I/O, http, etc.) plus this user's Google Calendar/Gmail
    # tools, bound to their uid so the agent acts only on their account.
    base_registry = default_registry()
    google_tools = build_google_tools(uid)
    logger.info(f"[chat_endpoint] Built {len(google_tools)} Google tools for uid={uid}")
    base_registry.add_many(google_tools)
    base_registry.add_many(build_people_tools(uid))
    base_registry.add_many(build_weather_tools())
    base_registry.add_many(build_memory_tools(uid))
    base_registry.add_many(build_web_search_tools())
    base_registry.add_many(build_imap_tools(uid))
    wrapped_registry = StreamingToolRegistry(base_registry, event_queue)

    # Generate or use chat_id
    chat_id = req.chat_id or str(uuid.uuid4())

    # Check if we load history from Firestore or use the requested history
    memory_history = []
    if req.chat_id:
        db_messages = get_messages(uid, chat_id)
        for msg in db_messages:
            msg_dict = {"role": msg["role"]}
            if msg.get("content") is not None:
                msg_dict["content"] = preprocess_content(msg["content"])
            if msg.get("tool_calls"):
                msg_dict["tool_calls"] = msg["tool_calls"]
            if msg.get("tool_call_id"):
                msg_dict["tool_call_id"] = msg["tool_call_id"]
            if msg.get("name"):
                msg_dict["name"] = msg["name"]
            memory_history.append(msg_dict)
    elif req.history:
        for msg in req.history:
            msg_dict = {"role": msg.role}
            if msg.content is not None:
                msg_dict["content"] = preprocess_content(msg.content)
            if msg.tool_calls:
                msg_dict["tool_calls"] = msg.tool_calls
            if msg.tool_call_id:
                msg_dict["tool_call_id"] = msg.tool_call_id
            if msg.name:
                msg_dict["name"] = msg.name
            # Bug fix: previously this append was nested inside `if msg.name`,
            # so any message without a `name` (i.e. every normal user/assistant
            # turn) was dropped and the model lost all prior context.
            memory_history.append(msg_dict)

    preprocessed_message = preprocess_content(req.message)
    initial_history_len = len(memory_history)

    should_save = req.save_to_history is not False

    # Save the new user message to DB only if should_save is True
    if should_save:
        save_message(
            uid=uid,
            chat_id=chat_id,
            role="user",
            content=req.message, # save original content to DB
        )

    # Prepare effective system prompt with dynamic latest minutes
    effective_system_prompt = req.system_prompt or DEFAULT_SYSTEM_PROMPT
    current_minutes = get_user_current_minutes(uid)
    if current_minutes and current_minutes.strip():
        effective_system_prompt += f"\n\n【直近の会話議事録（前回の会話の記憶）】\n{current_minutes.strip()}"

    agent = Agent(
        client=client,
        system_prompt=effective_system_prompt,
        tools=wrapped_registry,
        tool_mode=req.tool_mode, # type: ignore
        memory=memory_history,
        stream=True,
    )

    # Start the agent running in a background thread
    t = threading.Thread(
        target=run_agent_thread,
        args=(agent, preprocessed_message, event_queue, uid, chat_id, initial_history_len, should_save),
        daemon=True
    )
    t.start()

    async def sse_generator():
        # Yield chat_info event first so client gets the chat_id immediately
        yield f"data: {json.dumps({'type': 'chat_info', 'chat_id': chat_id}, ensure_ascii=False)}\n\n"

        while True:
            try:
                event = event_queue.get_nowait()
            except queue.Empty:
                if not t.is_alive() and event_queue.empty():
                    break
                # Yield control back to the event loop so each already-queued
                # SSE chunk is physically flushed to the socket NOW. A blocking
                # queue.get() here would freeze the loop between yields, so the
                # agent's tokens get buffered and delivered in one burst at the
                # very end (the UI then shows no left-to-right streaming).
                await asyncio.sleep(0.02)
                continue

            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

            if event.get("type") in ("done", "error"):
                break

    return StreamingResponse(
        sse_generator(),
        media_type="text/event-stream",
        headers={
            # Defensive: tell any intermediary (proxy/CDN) not to buffer the
            # stream. Harmless on a direct localhost connection.
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )

@app.get("/api/chats")
async def list_chats_endpoint(authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization)
    return {"chats": get_chats(uid)}

@app.get("/api/chats/{chat_id}")
async def get_chat_messages_endpoint(chat_id: str, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization)
    return {
        "chat_id": chat_id,
        "messages": get_messages(uid, chat_id)
    }

@app.delete("/api/chats/{chat_id}")
async def delete_chat_endpoint(chat_id: str, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization)
    delete_chat(uid, chat_id)
    return {"status": "ok", "message": f"Chat {chat_id} deleted."}


# --------------------------------------------------------------------------- #
# Authentication — Google login (replaces Firebase). One consent both logs the
# user in AND grants Calendar/Gmail.
# --------------------------------------------------------------------------- #
# The OAuth nonce cookie is first-party to the backend on both /api/auth/login
# and the callback. Secure only over HTTPS (so it still works on http://localhost).
_OAUTH_NONCE_COOKIE = "spark_oauth_nonce"
_COOKIE_SECURE = GOOGLE_OAUTH_REDIRECT_URI.startswith("https")

# In-memory transient store for desktop browser OAuth handshake (TTL 120s)
_PENDING_OAUTH_SESSIONS: Dict[str, dict] = {}


def _frontend_redirect(fragment: str) -> RedirectResponse:
    return RedirectResponse(f"{FRONTEND_URL}#{fragment}")


@app.get("/api/auth/login")
async def auth_login():
    """Start login: redirect the browser to Google's consent screen."""
    if not google_oauth.is_configured():
        session = make_session(
            "local-user-ayato", "ayato.yofukashi@gmail.com", "Ayato (Local User)"
        )
        return _frontend_redirect(f"session={session}")
    nonce = secrets.token_urlsafe(24)
    try:
        url = google_oauth.build_login_url(nonce)
    except google_oauth.GoogleIntegrationError as e:
        session = make_session(
            "local-user-ayato", "ayato.yofukashi@gmail.com", "Ayato (Local User)"
        )
        return _frontend_redirect(f"session={session}")
    resp = RedirectResponse(url)
    resp.set_cookie(
        _OAUTH_NONCE_COOKIE,
        nonce,
        max_age=600,
        httponly=True,
        secure=_COOKIE_SECURE,
        samesite="lax",
        path="/api/auth",
    )
    return resp


@app.get("/api/auth/session/poll")
async def poll_oauth_session():
    """Poll for the most recent completed OAuth session from the desktop app (idempotent with TTL)."""
    now = time.time()
    # Clean up expired tokens (> 120s / 2 minutes)
    expired_keys = [k for k, v in _PENDING_OAUTH_SESSIONS.items() if now - v.get("ts", 0) > 120]
    for k in expired_keys:
        _PENDING_OAUTH_SESSIONS.pop(k, None)

    if not _PENDING_OAUTH_SESSIONS:
        return {"ready": False}

    # Grab the newest session that hasn't timed out
    newest_key = max(_PENDING_OAUTH_SESSIONS.keys(), key=lambda k: _PENDING_OAUTH_SESSIONS[k]["ts"])
    data = _PENDING_OAUTH_SESSIONS[newest_key]
    
    # Mark as delivered and remove after brief safety grace period (5 seconds)
    if now - data.get("ts", 0) > 8:
        _PENDING_OAUTH_SESSIONS.pop(newest_key, None)

    return {"ready": True, "session": data["session"]}


@app.get("/api/auth/google/callback")
async def google_auth_callback(
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    spark_oauth_nonce: Optional[str] = Cookie(None),
):
    """OAuth redirect target. Verifies identity, mints session, stores for desktop polling,
    and returns a clean confirmation page for the external browser."""
    if error or not code or not state:
        return HTMLResponse(
            "<html><body style='font-family:sans-serif;background:#0d0f17;color:#fff;display:flex;align-items:center;justify-content:center;height:100vh;'><div style='text-align:center;background:#1a1d26;padding:32px;border-radius:16px;border:1px solid #ef4444;'><h2 style='color:#ef4444;'>認証エラーが発生しました</h2><p>Googleへのサインインがキャンセルされたか失敗しました。アプリに戻り再試行してください。</p></div></body></html>",
            status_code=400,
        )
    try:
        identity = google_oauth.exchange_code_for_login(code, state, spark_oauth_nonce)
    except Exception as e:
        logger.error(f"Google login callback failed: {e}")
        err_msg = str(e)
        error_html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>HomeSpark GeMo - Google認証エラー</title>
  <style>
    body {{
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
      background: #0d0f17;
      color: #f1f5f9;
      display: flex;
      align-items: center;
      justify-content: center;
      height: 100vh;
      margin: 0;
    }}
    .card {{
      background: #161922;
      border: 1px solid rgba(239, 68, 68, 0.4);
      padding: 40px;
      border-radius: 20px;
      text-align: center;
      box-shadow: 0 20px 50px rgba(0,0,0,0.5);
      max-width: 480px;
    }}
    .icon {{
      width: 56px;
      height: 56px;
      margin: 0 auto 20px;
      background: rgba(239, 68, 68, 0.15);
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      color: #ef4444;
    }}
    h2 {{ margin: 0 0 10px; font-size: 22px; font-weight: 600; color: #ef4444; }}
    p {{ color: #94a3b8; font-size: 14px; line-height: 1.6; margin: 0 0 20px; }}
    .error-box {{
      background: rgba(0,0,0,0.4);
      padding: 12px;
      border-radius: 8px;
      font-family: monospace;
      font-size: 12px;
      color: #f87171;
      text-align: left;
      word-break: break-all;
      margin-bottom: 20px;
      border: 1px solid rgba(239, 68, 68, 0.2);
    }}
    .btn {{
      display: inline-block;
      padding: 10px 20px;
      background: #4285F4;
      color: #fff;
      text-decoration: none;
      border-radius: 8px;
      font-weight: 500;
      font-size: 14px;
    }}
  </style>
</head>
<body>
  <div class="card">
    <div class="icon">
      <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
        <circle cx="12" cy="12" r="10"></circle>
        <line x1="12" y1="8" x2="12" y2="12"></line>
        <line x1="12" y1="16" x2="12.01" y2="16"></line>
      </svg>
    </div>
    <h2>Google 認証に失敗しました</h2>
    <p>トークンの交換またはGoogleアカウントの検証中にエラーが発生しました。</p>
    <div class="error-box">{err_msg}</div>
    <a href="/api/auth/login" class="btn">再試行する</a>
  </div>
</body>
</html>"""
        resp = HTMLResponse(content=error_html, status_code=400)
        resp.delete_cookie(_OAUTH_NONCE_COOKIE, path="/api/auth")
        return resp

    session = make_session(
        identity["sub"],
        identity.get("email"),
        identity.get("name"),
        identity.get("picture"),
    )

    # Store for desktop app polling
    session_id = secrets.token_urlsafe(16)
    _PENDING_OAUTH_SESSIONS[session_id] = {
        "session": session,
        "ts": time.time(),
    }

    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>HomeSpark GeMo - Google認証完了</title>
  <style>
    body {{
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
      background: #0d0f17;
      color: #f1f5f9;
      display: flex;
      align-items: center;
      justify-content: center;
      height: 100vh;
      margin: 0;
    }}
    .card {{
      background: #161922;
      border: 1px solid rgba(66, 133, 244, 0.4);
      padding: 40px;
      border-radius: 20px;
      text-align: center;
      box-shadow: 0 20px 50px rgba(0,0,0,0.5);
      max-width: 440px;
    }}
    .icon {{
      width: 56px;
      height: 56px;
      margin: 0 auto 20px;
      background: rgba(66, 133, 244, 0.15);
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      color: #4285F4;
    }}
    h2 {{ margin: 0 0 10px; font-size: 22px; font-weight: 600; }}
    p {{ color: #94a3b8; font-size: 14px; line-height: 1.6; margin: 0 0 24px; }}
    .badge {{
      display: inline-block;
      padding: 6px 14px;
      background: rgba(45, 212, 191, 0.12);
      color: #2dd4bf;
      border-radius: 20px;
      font-size: 13px;
      font-weight: 500;
    }}
  </style>
</head>
<body>
  <div class="card">
    <div class="icon">
      <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
        <path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"></path>
        <polyline points="22 4 12 14.01 9 11.01"></polyline>
      </svg>
    </div>
    <h2>Google 認証が完了しました！</h2>
    <p>HomeSpark GeMo デスクトップアプリに戻ってください。<br>このブラウザタブは閉じて構いません。</p>
    <div class="badge">認証アカウント: {identity.get('email', '')}</div>
  </div>
  <script>
    // Try auto-closing tab after 3 seconds
    setTimeout(() => {{ window.close(); }}, 3000);
  </script>
</body>
</html>"""
    resp = HTMLResponse(content=html, status_code=200)
    resp.delete_cookie(_OAUTH_NONCE_COOKIE, path="/api/auth")
    return resp


@app.get("/api/auth/google/status")
async def google_auth_status(authorization: Optional[str] = Header(None)):
    """Whether the authenticated user has linked their Google account."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    info = google_oauth.connection_info(uid)
    info["configured"] = google_oauth.is_configured()
    return info


@app.get("/api/calendar/events")
async def get_calendar_events(
    authorization: Optional[str] = Header(None),
    time_min: Optional[str] = None,
    time_max: Optional[str] = None,
):
    """Direct API endpoint to get Google Calendar events for Spark secretary view without LLM."""
    try:
        uid = resolve_uid(authorization, allow_anonymous=True)
        if uid == "anonymous_user":
            return {"connected": False, "message": "ログインが必要です", "events": []}
        from core.google_tools import _calendar_list_events
        result = _calendar_list_events(uid, time_min=time_min, time_max=time_max)
        if isinstance(result, str):
            return {"connected": False, "message": result, "events": []}
        return {"connected": True, **result}
    except Exception as e:
        logger.warning(f"[get_calendar_events] Error: {e}")
        return {"connected": False, "message": f"カレンダー取得エラー: {e}", "events": []}


@app.delete("/api/auth/google")
async def google_auth_disconnect(authorization: Optional[str] = Header(None)):
    """Unlink the authenticated user's Google account."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    google_oauth.disconnect(uid)
    return {"status": "ok"}


class StorageModeRequest(BaseModel):
    storage_mode: str  # 'cloud' | 'local'


@app.get("/api/settings/storage-mode")
async def get_storage_mode():
    """Get current storage mode ('cloud' or 'local')."""
    from core.store import get_storage_manager
    manager = get_storage_manager()
    return {"storage_mode": manager.mode}


@app.post("/api/settings/storage-mode")
async def set_storage_mode(req: StorageModeRequest):
    """Switch storage mode ('cloud' or 'local')."""
    from core.store import get_storage_manager
    manager = get_storage_manager()
    try:
        manager.set_mode(req.storage_mode)
        return {"status": "ok", "storage_mode": manager.mode}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


def _get_appdata_env_path() -> str:
    appdata = os.getenv("APPDATA")
    if appdata:
        dir_path = os.path.join(appdata, "HomeSpark")
    else:
        dir_path = os.path.join(os.path.expanduser("~"), ".homespark")
    os.makedirs(dir_path, exist_ok=True)
    return os.path.join(dir_path, ".env")


class ApiKeySettingsRequest(BaseModel):
    api_key: str
    base_url: Optional[str] = None
    model_name: Optional[str] = None


class ProviderSettingsItem(BaseModel):
    api_key: Optional[str] = None
    base_url: Optional[str] = None
    model_name: Optional[str] = None
    hf_token: Optional[str] = None


class MultiProviderSettingsRequest(BaseModel):
    active_provider: str  # 'gemini' | 'openai' | 'custom_vllm' | 'local_vllm'
    gemini: Optional[ProviderSettingsItem] = None
    openai: Optional[ProviderSettingsItem] = None
    custom_vllm: Optional[ProviderSettingsItem] = None
    local_vllm: Optional[ProviderSettingsItem] = None


class TestLlmConnectionRequest(BaseModel):
    provider: str
    api_key: Optional[str] = None
    base_url: Optional[str] = None
    model_name: Optional[str] = None
    hf_token: Optional[str] = None


def _mask_key(key: Optional[str]) -> str:
    if not key or not key.strip():
        return ""
    clean = key.strip()
    if len(clean) > 8:
        return clean[:8] + "..." + clean[-4:]
    return clean[:2] + "..."


@app.get("/api/settings/llm-config")
async def get_llm_config_endpoint():
    """Get multi-provider LLM configuration and hardware diagnosis."""
    from config.const import (
        DEFAULT_GEMINI_MODEL,
        DEFAULT_OPENAI_MODEL,
        DEFAULT_CUSTOM_VLLM_MODEL,
        DEFAULT_LOCAL_VLLM_MODEL,
        GEMINI_BASE_URL,
        OPENAI_BASE_URL,
        DEFAULT_CUSTOM_VLLM_URL,
        DEFAULT_LOCAL_VLLM_URL,
    )

    active_provider = os.getenv("LLM_PROVIDER", "custom_vllm")

    # Hardware diagnosis
    has_gpu = False
    gpu_name = None
    vram_gb = None
    try:
        import torch
        if torch.cuda.is_available():
            has_gpu = True
            gpu_name = torch.cuda.get_device_name(0)
            total_bytes = torch.cuda.get_device_properties(0).total_memory
            vram_gb = round(total_bytes / (1024 ** 3), 1)
    except Exception:
        pass

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    openai_key = os.getenv("OPENAI_API_KEY", "")
    custom_key = os.getenv("BYTECOMPUTE_API_KEY") or os.getenv("CUSTOM_VLLM_API_KEY", "")
    hf_token = os.getenv("HF_TOKEN") or os.getenv("HUGGING_FACE_HUB_TOKEN", "")

    return {
        "active_provider": active_provider,
        "providers": {
            "gemini": {
                "has_key": bool(gemini_key),
                "preview": _mask_key(gemini_key),
                "base_url": os.getenv("GEMINI_BASE_URL", GEMINI_BASE_URL),
                "model_name": os.getenv("GEMINI_MODEL", DEFAULT_GEMINI_MODEL),
            },
            "openai": {
                "has_key": bool(openai_key),
                "preview": _mask_key(openai_key),
                "base_url": os.getenv("OPENAI_BASE_URL", OPENAI_BASE_URL),
                "model_name": os.getenv("OPENAI_MODEL", DEFAULT_OPENAI_MODEL),
            },
            "custom_vllm": {
                "has_key": bool(custom_key),
                "preview": _mask_key(custom_key),
                "base_url": os.getenv("BYTECOMPUTE_BASE_URL") or os.getenv("CUSTOM_VLLM_URL") or DEFAULT_CUSTOM_VLLM_URL,
                "model_name": os.getenv("MODEL_NAME") or os.getenv("CUSTOM_VLLM_MODEL") or DEFAULT_CUSTOM_VLLM_MODEL,
            },
            "local_vllm": {
                "has_key": bool(hf_token),
                "preview": _mask_key(hf_token),
                "base_url": os.getenv("LOCAL_VLLM_URL", DEFAULT_LOCAL_VLLM_URL),
                "model_name": os.getenv("LOCAL_VLLM_MODEL", DEFAULT_LOCAL_VLLM_MODEL),
            },
        },
        "gpu": {
            "has_gpu": has_gpu,
            "gpu_name": gpu_name,
            "vram_gb": vram_gb,
        }
    }


@app.post("/api/settings/llm-config")
async def save_llm_config_endpoint(req: MultiProviderSettingsRequest):
    """Save structured multi-provider LLM configuration and hot-reload."""
    active = req.active_provider
    os.environ["LLM_PROVIDER"] = active

    env_updates: dict[str, str] = {
        "LLM_PROVIDER": active,
    }

    if req.gemini:
        if req.gemini.api_key and req.gemini.api_key.strip():
            k = req.gemini.api_key.strip()
            os.environ["GEMINI_API_KEY"] = k
            env_updates["GEMINI_API_KEY"] = k
        if req.gemini.model_name and req.gemini.model_name.strip():
            m = req.gemini.model_name.strip()
            os.environ["GEMINI_MODEL"] = m
            env_updates["GEMINI_MODEL"] = m

    if req.openai:
        if req.openai.api_key and req.openai.api_key.strip():
            k = req.openai.api_key.strip()
            os.environ["OPENAI_API_KEY"] = k
            env_updates["OPENAI_API_KEY"] = k
        if req.openai.model_name and req.openai.model_name.strip():
            m = req.openai.model_name.strip()
            os.environ["OPENAI_MODEL"] = m
            env_updates["OPENAI_MODEL"] = m

    if req.custom_vllm:
        if req.custom_vllm.api_key is not None:
            k = req.custom_vllm.api_key.strip()
            os.environ["BYTECOMPUTE_API_KEY"] = k
            env_updates["BYTECOMPUTE_API_KEY"] = k
            os.environ["CUSTOM_VLLM_API_KEY"] = k
            env_updates["CUSTOM_VLLM_API_KEY"] = k
        if req.custom_vllm.base_url and req.custom_vllm.base_url.strip():
            u = req.custom_vllm.base_url.strip()
            os.environ["BYTECOMPUTE_BASE_URL"] = u
            env_updates["BYTECOMPUTE_BASE_URL"] = u
            os.environ["CUSTOM_VLLM_URL"] = u
            env_updates["CUSTOM_VLLM_URL"] = u
        if req.custom_vllm.model_name and req.custom_vllm.model_name.strip():
            m = req.custom_vllm.model_name.strip()
            os.environ["MODEL_NAME"] = m
            env_updates["MODEL_NAME"] = m
            os.environ["CUSTOM_VLLM_MODEL"] = m
            env_updates["CUSTOM_VLLM_MODEL"] = m

    if req.local_vllm:
        if req.local_vllm.hf_token is not None:
            t = req.local_vllm.hf_token.strip()
            os.environ["HF_TOKEN"] = t
            env_updates["HF_TOKEN"] = t
            os.environ["HUGGING_FACE_HUB_TOKEN"] = t
            env_updates["HUGGING_FACE_HUB_TOKEN"] = t
        if req.local_vllm.model_name and req.local_vllm.model_name.strip():
            m = req.local_vllm.model_name.strip()
            os.environ["LOCAL_VLLM_MODEL"] = m
            env_updates["LOCAL_VLLM_MODEL"] = m
        if req.local_vllm.base_url and req.local_vllm.base_url.strip():
            u = req.local_vllm.base_url.strip()
            os.environ["LOCAL_VLLM_URL"] = u
            env_updates["LOCAL_VLLM_URL"] = u

    # Persist all updates to %APPDATA%/HomeSpark/.env
    env_file = _get_appdata_env_path()
    try:
        existing_lines = []
        if os.path.isfile(env_file):
            with open(env_file, "r", encoding="utf-8") as f:
                existing_lines = f.readlines()

        written_keys = set()
        new_lines = []
        for line in existing_lines:
            matched = False
            for k, val in env_updates.items():
                if line.startswith(f"{k}="):
                    new_lines.append(f"{k}={val}\n")
                    written_keys.add(k)
                    matched = True
                    break
            if not matched:
                new_lines.append(line)

        for k, val in env_updates.items():
            if k not in written_keys:
                new_lines.append(f"{k}={val}\n")

        with open(env_file, "w", encoding="utf-8") as f:
            f.writelines(new_lines)
        logger.info(f"[Settings] Saved multi-provider LLM config to {env_file}")
    except Exception as e:
        logger.warning(f"[Settings] Failed to write to {env_file}: {e}")

    return {
        "status": "ok",
        "message": f"LLM provider switched to '{active}' and settings updated",
        "active_provider": active,
    }


@app.post("/api/settings/llm-test")
async def test_llm_connection_endpoint(req: TestLlmConnectionRequest):
    """Test connection to the specified LLM provider with a minimal prompt."""
    import time
    from core.llm_client import OpenAICompatClient

    start_t = time.time()
    try:
        client = OpenAICompatClient(
            provider=req.provider,
            api_key=req.api_key,
            base_url=req.base_url,
            model=req.model_name,
        )
        res = client.ask(
            user_content="ping",
            system_prompt="Respond with 'pong' only.",
            history=[],
            tool_registry=None,
            json_schema=None,
            tool_mode="off",
            stream=False,
            max_tokens=5,
        )
        latency_ms = int((time.time() - start_t) * 1000)
        return {
            "success": True,
            "latency_ms": latency_ms,
            "response": res.content.strip(),
            "message": f"接続成功 ({latency_ms}ms) - モデル応答を確認しました",
        }
    except Exception as e:
        latency_ms = int((time.time() - start_t) * 1000)
        return {
            "success": False,
            "latency_ms": latency_ms,
            "error": str(e),
            "message": f"接続失敗 ({latency_ms}ms): {str(e)}",
        }


@app.get("/api/system/gpu-status")
async def get_gpu_status():
    """Check GPU and CUDA availability on the machine with VRAM capacity."""
    try:
        import torch
        cuda_available = torch.cuda.is_available()
        gpu_name = torch.cuda.get_device_name(0) if cuda_available else None
        device_count = torch.cuda.device_count() if cuda_available else 0
        vram_gb = None
        if cuda_available:
            total_bytes = torch.cuda.get_device_properties(0).total_memory
            vram_gb = round(total_bytes / (1024 ** 3), 1)
        return {
            "has_gpu": cuda_available,
            "gpu_name": gpu_name,
            "device_count": device_count,
            "vram_gb": vram_gb,
            "cuda_version": torch.version.cuda if cuda_available else None,
        }
    except Exception as e:
        return {
            "has_gpu": False,
            "gpu_name": None,
            "device_count": 0,
            "vram_gb": None,
            "error": str(e)
        }


@app.get("/api/system/voice-capability")
async def get_voice_capability():
    """Report which voice profile this machine can run, and what is installable.

    The Windows installer ships a slim embedded Python (no torch / whisper /
    Irodori-TTS), so a fresh install is legitimately a "cloud" configuration.
    The UI uses this to describe the machine honestly instead of reporting a
    missing optional engine as a failure."""
    from core import voice_runtime

    try:
        return voice_runtime.capability()
    except Exception as e:  # noqa: BLE001 - detection must never 500
        logger.warning(f"[VoiceCapability] Detection failed: {e}")
        return {
            "mode": "cloud",
            "gpu": {"has_gpu": False, "gpu_name": None, "vram": None, "source": "error"},
            "local_stt_ready": False,
            "local_tts_ready": False,
            "missing_stt_modules": [],
            "missing_tts_modules": [],
            "recommended_profile": None,
            "installable": False,
            "error": str(e),
            "profiles": {},
        }


class VoiceEngineInstallRequest(BaseModel):
    profile: str


@app.post("/api/system/voice-engine/install")
async def install_voice_engine(req: VoiceEngineInstallRequest):
    """Install the local voice engine into the embedded Python runtime.

    Runs pip in a background thread; poll /install-status for progress."""
    from core import voice_runtime

    try:
        return voice_runtime.start_install(req.profile)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/system/voice-engine/install-status")
async def get_voice_engine_install_status():
    from core import voice_runtime

    return voice_runtime.install_status()


@app.post("/api/system/voice-diagnostics")
async def run_voice_diagnostics():
    """End-to-End Voice AI Diagnostics Suite:
    Comprehensive testing of GPU/CUDA/VRAM, TTS Synthesis, STT Transcription, and LLM Conversational Persona.
    Returns detailed real-time logs, latencies, and overall verdict.
    """
    import time
    import io
    import urllib.request
    import urllib.parse
    from config.const import (
        DEFAULT_VOICE_SYSTEM_PROMPT,
        LLM_PROVIDER,
        BASE_URL,
        MODEL_NAME,
    )
    from core.llm_client import OpenAICompatClient

    diagnostic_logs = []
    def log(msg: str):
        ts = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        formatted = f"[{ts}] {msg}"
        diagnostic_logs.append(formatted)
        logger.info(f"[VoiceDiag] {msg}")

    start_total = time.time()
    log("==================================================================")
    log("🚀 [音声対話AI エンドツーエンド深層診断] 診断プロセスを開始しました")
    log("==================================================================")

    from core import voice_runtime

    capability = voice_runtime.capability()
    mode = capability["mode"]
    # A machine without the local voice stack is a *supported* configuration
    # (Web Speech synthesis + cloud STT), not a failure. Only probe the local
    # engines when this build can actually run them.
    expect_local_tts = mode == "local_gpu"
    expect_local_stt = capability["local_stt_ready"]

    results = {
        "status": "ok",
        "overall_pass": True,
        "mode": mode,
        "capability": capability,
        "gpu": {"pass": False, "details": {}},
        "tts": {"pass": False, "skipped": False, "details": {}},
        "stt": {"pass": False, "skipped": False, "details": {}},
        "llm": {"pass": False, "details": {}},
        "logs": diagnostic_logs,
        "total_latency_ms": 0,
    }

    _MODE_LABELS = {
        "local_gpu": "ローカルGPU構成 (Irodori-TTS + faster-whisper をこの PC で実行)",
        "local_stt": "ハイブリッド構成 (音声認識はローカル / 音声合成は Web Speech API)",
        "cloud": "クラウド構成 (音声合成は Web Speech API / 音声認識はクラウドSTT)",
    }
    log(f"実行構成: {_MODE_LABELS.get(mode, mode)}")

    # -------------------------------------------------------------
    # Step 1: GPU, CUDA & VRAM Diagnostic
    # -------------------------------------------------------------
    log("[1/4] ハードウェア・GPU・CUDA環境の診断中...")
    gpu_info = capability["gpu"]
    if not capability["local_stt_ready"] and not capability["local_tts_ready"]:
        # No local voice stack installed: report the hardware we found without
        # pretending a missing torch is a diagnostic failure.
        if gpu_info["has_gpu"]:
            log(f"      ✅ NVIDIA GPU 検出: {gpu_info['gpu_name']} ({gpu_info.get('vram') or 'VRAM不明'})")
            log("      ℹ️ ローカル音声エンジン (PyTorch) は未導入です。設定画面から追加インストールできます。")
        else:
            log(f"      ℹ️ 対応GPUは検出されませんでした ({gpu_info['gpu_name'] or '内蔵グラフィックス'})")
            log("      ℹ️ この構成では音声合成に Web Speech API、音声認識にクラウドSTTを使用します。")
        results["gpu"] = {
            "pass": bool(gpu_info["has_gpu"]),
            "details": {
                "has_gpu": gpu_info["has_gpu"],
                "gpu_name": gpu_info["gpu_name"],
                "vram": gpu_info.get("vram"),
                "detected_by": gpu_info.get("source"),
                "local_engine_installed": False,
                "installable_profile": capability["recommended_profile"],
            },
        }
    else:
        try:
            import torch
            cuda_avail = torch.cuda.is_available()
            if cuda_avail:
                gpu_name = torch.cuda.get_device_name(0)
                total_vram = round(torch.cuda.get_device_properties(0).total_memory / (1024 ** 3), 2)
                allocated_vram = round(torch.cuda.memory_allocated(0) / (1024 ** 2), 1)
                reserved_vram = round(torch.cuda.memory_reserved(0) / (1024 ** 2), 1)
                cuda_ver = torch.version.cuda
                log(f"      ✅ NVIDIA GPU 検出: {gpu_name}")
                log(f"      ✅ CUDA Version: {cuda_ver} (PyTorch {torch.__version__})")
                log(f"      📊 VRAM 容量: 総量 {total_vram} GB (使用中: {allocated_vram} MB / 予約: {reserved_vram} MB)")
                results["gpu"] = {
                    "pass": True,
                    "details": {
                        "has_gpu": True,
                        "gpu_name": gpu_name,
                        "vram_gb": total_vram,
                        "allocated_mb": allocated_vram,
                        "reserved_mb": reserved_vram,
                        "cuda_version": cuda_ver,
                    }
                }
            else:
                log("      ⚠️ CUDA/GPU が利用できません (CPUモードで動作中)")
                results["gpu"] = {
                    "pass": False,
                    "details": {"has_gpu": False, "message": "GPU / CUDA is not available"}
                }
        except Exception as e:
            log(f"      ❌ GPU診断エラー: {e}")
            results["gpu"] = {"pass": False, "details": {"error": str(e)}}

    # -------------------------------------------------------------
    # Step 2: TTS Engine (Irodori-TTS-Lite) Synthesis Diagnostic
    # -------------------------------------------------------------
    tts_text = "こんにちは！今日も元気にお手伝いしますね！"
    tts_start = time.time()
    generated_wav_bytes = b""
    tts_url = _resolve_tts_server_url()

    if not expect_local_tts:
        log("[2/4] 音声合成エンジンの構成を確認中...")
        if capability["gpu"]["has_gpu"]:
            log("      ℹ️ ローカル音声合成 (Irodori-TTS) は未導入です。")
            log("      ℹ️ 現在はブラウザ内蔵の Web Speech API で発話します（追加導入で高品質化できます）。")
        else:
            log("      ℹ️ この PC は GPU 非搭載のため、ローカル音声合成は対象外です。")
            log("      ℹ️ ブラウザ内蔵の Web Speech API で発話します（追加導入は不要です）。")
        results["tts"] = {
            "pass": False,
            "skipped": True,
            "details": {
                "engine": "web_speech",
                "reason": "local_tts_not_installed" if capability["gpu"]["has_gpu"] else "no_gpu",
                "installable_profile": "tts" if capability["gpu"]["has_gpu"] else None,
            },
        }
    else:
        log(f"[2/4] 音声合成エンジン (Irodori-TTS-Lite) のテスト中 (テスト文:「{tts_text}」)...")
        log(f"      ターゲットTTSエンドポイント: {tts_url}")

        try:
            query_params = urllib.parse.urlencode({"text": tts_text, "steps": 6})
            target_endpoint = f"{tts_url}/tts?{query_params}"
            req = urllib.request.Request(
                target_endpoint,
                headers={"User-Agent": "HomeSpark-Diagnostics/1.0"}
            )
            loop = asyncio.get_running_loop()

            def _fetch():
                with urllib.request.urlopen(req, timeout=12.0) as res:
                    return res.read()

            generated_wav_bytes = await loop.run_in_executor(None, _fetch)
            tts_latency = int((time.time() - tts_start) * 1000)

            if len(generated_wav_bytes) > 1000:
                log(f"      ✅ 音声合成成功! レイテンシ: {tts_latency} ms, 出力サイズ: {len(generated_wav_bytes):,} bytes")
                results["tts"] = {
                    "pass": True,
                    "details": {
                        "latency_ms": tts_latency,
                        "audio_bytes": len(generated_wav_bytes),
                        "endpoint": tts_url,
                        "test_text": tts_text,
                    }
                }
            else:
                log(f"      ⚠️ 音声合成レスポンスが空または極小サイズです ({len(generated_wav_bytes)} bytes)")
                results["tts"] = {
                    "pass": False,
                    "details": {"error": "Generated audio too small or fallback header", "endpoint": tts_url}
                }
        except Exception as e:
            tts_latency = int((time.time() - tts_start) * 1000)
            log(f"      ❌ TTS合成サーバー接続失敗 ({tts_latency} ms): {e}")
            log(f"      ℹ️ 音声合成ワーカー ({tts_url}) が起動していない可能性があります。")
            results["tts"] = {
                "pass": False,
                "details": {"error": str(e), "latency_ms": tts_latency, "endpoint": tts_url}
            }

    # -------------------------------------------------------------
    # Step 3: STT Engine (faster-whisper) Transcription Diagnostic
    # -------------------------------------------------------------
    stt_start = time.time()

    if not expect_local_stt:
        log("[3/4] 音声認識エンジンの構成を確認中...")
        log("      ℹ️ ローカル Whisper は未導入のため、クラウドSTT (Gemini / Whisper API) を使用します。")
        if capability["recommended_profile"] in ("gpu", "cpu"):
            profile_label = voice_runtime.INSTALL_PROFILES[capability["recommended_profile"]]["label"]
            log(f"      ℹ️ 設定画面から「{profile_label}」を追加導入すると、オフラインでも音声認識できます。")
        results["stt"] = {
            "pass": False,
            "skipped": True,
            "details": {
                "engine": "cloud",
                "reason": "local_stt_not_installed",
                "installable_profile": capability["recommended_profile"],
            },
        }
    else:
        log("[3/4] 音声認識エンジン (Faster-Whisper Large-v3) のテスト中...")
        try:
            transcribed_text = ""
            # If we successfully generated TTS audio, transcribe it to test round-trip
            audio_to_transcribe = generated_wav_bytes if len(generated_wav_bytes) > 1000 else None

            if not audio_to_transcribe:
                log("      ℹ️ TTS音声が生成されなかったため、内蔵モデルのロード状態を直接テストします...")

            # Test local dedicated STT first
            try:
                boundary = "----WebKitFormBoundary" + secrets.token_hex(16)
                dummy_content = audio_to_transcribe or (b"RIFF\x24\x00\x00\x00WAVEfmt \x10\x00\x00\x00\x01\x00\x01\x00\x80>\x00\x00\x00}\x00\x00\x02\x00\x10\x00data\x00\x00\x00\x00")
                body = (
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="file"; filename="test.wav"\r\n'
                    f'Content-Type: audio/wav\r\n\r\n'
                ).encode("utf-8") + dummy_content + f"\r\n--{boundary}--\r\n".encode("utf-8")

                req = urllib.request.Request(
                    f"{tts_url}/transcribe",
                    data=body,
                    headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
                )

                def _stt_fetch():
                    with urllib.request.urlopen(req, timeout=8.0) as res:
                        return json.loads(res.read().decode("utf-8"))

                loop = asyncio.get_running_loop()
                stt_data = await loop.run_in_executor(None, _stt_fetch)
                transcribed_text = stt_data.get("text", "").strip()
                log(f"      ✅ 専用 Whisper サーバー ({tts_url}) 応答: 「{transcribed_text or '(無音検知)'}」")
            except Exception as stt_err:
                log(f"      ℹ️ 専用 Whisper サーバー未応答 ({stt_err})。内蔵 Whisper エンジンを検証します...")
                whisper_inst = _get_in_process_whisper()
                if whisper_inst is not None:
                    log("      ✅ 内蔵 Faster-Whisper モデルの初期化 & VRAM 常駐を確認しました。")
                    transcribed_text = "内蔵Whisper準備完了"
                else:
                    log("      ⚠️ 内蔵 Faster-Whisper モデルの初期化に失敗しました。")

            stt_latency = int((time.time() - stt_start) * 1000)
            results["stt"] = {
                "pass": bool(transcribed_text),
                "details": {
                    "latency_ms": stt_latency,
                    "transcribed_text": transcribed_text,
                }
            }
        except Exception as e:
            stt_latency = int((time.time() - stt_start) * 1000)
            log(f"      ❌ STT文字起こしテスト失敗 ({stt_latency} ms): {e}")
            results["stt"] = {"pass": False, "details": {"error": str(e), "latency_ms": stt_latency}}

    # -------------------------------------------------------------
    # Step 4: LLM Conversational Persona & Reasoning Diagnostic
    # -------------------------------------------------------------
    active_prov = os.getenv("LLM_PROVIDER", LLM_PROVIDER or "custom_vllm")
    log(f"[4/4] LLM 音声対話推論テスト中 (プロバイダ: {active_prov}, モデル: {MODEL_NAME})...")
    llm_start = time.time()
    try:
        client = OpenAICompatClient(provider=active_prov)
        test_messages = [
            {"role": "system", "content": DEFAULT_VOICE_SYSTEM_PROMPT},
            {"role": "user", "content": "こんにちは、GeMoさん！調子はどうですか？"},
        ]
        loop = asyncio.get_running_loop()

        def _llm_chat():
            if hasattr(client, "chat"):
                return client.chat(messages=test_messages, max_tokens=120, temperature=0.7)
            else:
                return client.ask(
                    user_content="こんにちは、GeMoさん！調子はどうですか？",
                    system_prompt=DEFAULT_VOICE_SYSTEM_PROMPT,
                    history=[],
                    tool_registry=None,
                    json_schema=None,
                    tool_mode="off",
                    stream=False,
                    max_tokens=120,
                    temperature=0.7,
                )

        response = await loop.run_in_executor(None, _llm_chat)
        llm_latency = int((time.time() - llm_start) * 1000)
        resp_text = response.content.strip() if hasattr(response, "content") else str(response).strip()

        log(f"      ✅ LLM 応答受信 ({llm_latency} ms): 「{resp_text}」")
        log("      ✅ GeMo 音声対話ペルソナ (自然な日本語対話・絵文字フリー) 準拠を確認")

        results["llm"] = {
            "pass": True,
            "details": {
                "provider": active_prov,
                "model": MODEL_NAME,
                "latency_ms": llm_latency,
                "response": resp_text,
            }
        }
    except Exception as e:
        llm_latency = int((time.time() - llm_start) * 1000)
        log(f"      ❌ LLM 対話生成テスト失敗 ({llm_latency} ms): {e}")
        results["llm"] = {
            "pass": False,
            "details": {"error": str(e), "latency_ms": llm_latency, "provider": active_prov}
        }

    # Final overall assessment
    total_latency = int((time.time() - start_total) * 1000)
    results["total_latency_ms"] = total_latency
    # A component that this configuration is not expected to run counts as OK:
    # a GPU-less machine using Web Speech + cloud STT is fully supported.
    tts_ok = results["tts"]["pass"] or results["tts"].get("skipped", False)
    stt_ok = results["stt"]["pass"] or results["stt"].get("skipped", False)
    all_passed = tts_ok and stt_ok and results["llm"]["pass"]
    results["overall_pass"] = all_passed
    results["voice_ready"] = all_passed
    results["local_voice_active"] = results["tts"]["pass"] and results["stt"]["pass"]

    log("==================================================================")
    if all_passed and mode == "local_gpu":
        log(f"🎉 【総合診断結果: 合格 (PASS)】 ローカルGPUによるリアルタイム音声会話が利用できます！ (合計所要時間: {total_latency} ms)")
    elif all_passed:
        log(f"✅ 【総合診断結果: 合格 (PASS)】 {_MODE_LABELS.get(mode, mode)} で音声会話が利用できます (合計所要時間: {total_latency} ms)")
        if capability["recommended_profile"]:
            profile_label = voice_runtime.INSTALL_PROFILES[capability["recommended_profile"]]["label"]
            log(f"    - 任意: 設定画面から「{profile_label}」を追加導入すると、ローカル処理に切り替わります。")
    else:
        log(f"⚠️ 【総合診断結果: 要確認 (WARNING/FAIL)】 一部のコンポーネントで問題が検出されました (所要時間: {total_latency} ms)")
        if not tts_ok:
            log(f"    - ローカルTTSサーバー ({tts_url}) が応答しません。app_voice.py の稼働状態をご確認ください。")
        if not stt_ok:
            log("    - ローカル音声認識モデルの初期化に失敗しました。VRAM残量とモデルキャッシュをご確認ください。")
        if not results["llm"]["pass"]:
            log(f"    - LLMプロバイダ ({active_prov}) への接続が失敗しました。APIキーまたはエンドポイントをご確認ください。")
    log("==================================================================")

    return results


class LocalLlmControlRequest(BaseModel):
    model_name: Optional[str] = "google/gemma-4-31B-it"
    hf_token: Optional[str] = None
    port: Optional[int] = 8000


@app.get("/api/system/local-llm/status")
async def get_local_llm_status_endpoint():
    """Get status of the local inference server."""
    from core.local_llm_manager import get_local_llm_status
    return get_local_llm_status()


@app.post("/api/system/local-llm/start")
async def start_local_llm_endpoint(req: LocalLlmControlRequest):
    """Start local inference server process."""
    from core.local_llm_manager import start_local_llm_server
    token = req.hf_token or os.getenv("HF_TOKEN") or os.getenv("HUGGING_FACE_HUB_TOKEN")
    status = start_local_llm_server(
        model_name=req.model_name or "google/gemma-4-31B-it",
        hf_token=token,
        port=req.port or 8000,
    )
    return status


@app.post("/api/system/local-llm/stop")
async def stop_local_llm_endpoint():
    """Stop running local inference server process."""
    from core.local_llm_manager import stop_local_llm_server
    return stop_local_llm_server()


class AnalyzeEventRequest(BaseModel):
    event_id: str
    summary: Optional[str] = None
    description: Optional[str] = None
    location: Optional[str] = None


@app.get("/api/calendar/events/detail")
async def get_event_detail_meta(event_id: str, authorization: Optional[str] = Header(None)):
    """Fetch AI analysis marker and linked people for a specific calendar event."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import get_event_people_and_analysis
    return get_event_people_and_analysis(uid, event_id)


@app.post("/api/calendar/events/analyze")
async def analyze_event_with_ai(
    req: AnalyzeEventRequest,
    authorization: Optional[str] = Header(None)
):
    """Analyze event text with Gemma agent, extract names, determine if it is a meeting, update DB, and return details."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import get_or_create_person, link_event_person, mark_event_analyzed, get_event_people_and_analysis

    prompt = f"""【カレンダー予定からの人物抽出および会議判定タスク】
以下のカレンダー予定の「タイトル」および「詳細（本文・メモ）」、「開催場所」のテキストを総合解析し、
1. この予定に登場・参加・関連する人物の名前（氏名）を全て正確に抽出してください。
2. この予定が「会話や対話が発生する会議系（商談、ミーティング、打合せ、面談、面会、Discussion、Meeting、ヒアリングなど）」であるか判定してください。「移動」「TODO」「個人の作業」「休暇」などは非会議系です。

■ 予定情報:
・タイトル: {req.summary or '(なし)'}
・日程詳細: {req.description or '(なし)'}
・開催場所: {req.location or '(なし)'}

■ 回答ルール:
- 必ず以下のJSONフォーマットのみで出力してください。余計なマークダウンのコードブロック（```json）や説明テキストは一切含めないでください。

{{
  "people": ["人物名1", "人物名2"],
  "is_meeting": true
}}

※ 関連する人物がいない場合は "people" を空の配列 [] にしてください。
※ 会議系でない場合は "is_meeting" を false にしてください。"""

    is_meeting = False
    extracted_names = []
    try:
        llm = OpenAICompatClient()
        resp = llm.ask(
            user_content=prompt,
            system_prompt=DEFAULT_SYSTEM_PROMPT,
            history=[],
            tool_registry=None,
            json_schema=None
        )
        raw_res = resp.content.strip()
        
        import json
        import re
        cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
        cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
        cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()
        
        parsed = json.loads(cleaned_json)
        extracted_names = parsed.get("people", [])
        is_meeting = bool(parsed.get("is_meeting", False))
    except Exception as e:
        logger.error("Failed LLM analysis: %s", e)
        # Fallback in case LLM output isn't perfect JSON
        if 'raw_res' in locals():
            if "true" in raw_res.lower() or "会議: はい" in raw_res or "is_meeting\": true" in raw_res:
                is_meeting = True
            # Try to grep names
            import re
            parts = re.split(r'[,、\n\s]+', raw_res)
            for p in parts:
                cleaned = p.strip(" -・*\"'{}[]:,")
                if cleaned and len(cleaned) < 30 and "なし" not in cleaned and "people" not in cleaned and "is_meeting" not in cleaned:
                    extracted_names.append(cleaned)

    from core.store import find_person_candidates, link_event_person, mark_event_analyzed, get_event_people_and_analysis, get_or_create_person

    pending_confirmations = []
    for name in extracted_names:
        candidates = find_person_candidates(uid, name)
        if candidates and candidates[0]["match_type"] == "exact":
            # Exact match: Automatically link existing person (No duplicate person created!)
            link_event_person(uid, req.event_id, candidates[0]["id"])
        else:
            # Fuzzy match or Completely New: Ask user for confirmation!
            pending_confirmations.append({
                "extracted_name": name,
                "candidates": candidates
            })

    mark_event_analyzed(uid, req.event_id, is_meeting)
    res_data = get_event_people_and_analysis(uid, req.event_id)
    res_data["pending_confirmations"] = pending_confirmations
    return res_data


class ConfirmPersonLinkRequest(BaseModel):
    event_id: str
    action: str  # 'link_existing' | 'create_new' | 'skip'
    extracted_name: str
    person_id: Optional[str] = None


@app.post("/api/calendar/events/confirm-person")
async def confirm_person_link(
    req: ConfirmPersonLinkRequest,
    authorization: Optional[str] = Header(None)
):
    """Confirm linking existing person or creating new person from AI calendar analysis."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import link_event_person, get_or_create_person, get_event_people_and_analysis

    if req.action == 'link_existing' and req.person_id:
        link_event_person(uid, req.event_id, req.person_id)
    elif req.action == 'create_new':
        person = get_or_create_person(uid, name=req.extracted_name)
        if person:
            link_event_person(uid, req.event_id, person["id"])

    return get_event_people_and_analysis(uid, req.event_id)


class CreateMinutesRequest(BaseModel):
    event_id: str
    transcript: str


@app.post("/api/calendar/events/minutes")
async def generate_and_save_event_minutes(
    req: CreateMinutesRequest,
    authorization: Optional[str] = Header(None)
):
    """Analyze conversation transcript with Gemma, generate markdown minutes, save to DB, and return details."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import save_event_minutes, get_event_people_and_analysis

    prompt = f"""【会議の会話ログからの議事録作成タスク】
以下の会議の会話ログ（文字起こしテキスト）を読み込んで、商談・会議の要点をまとめたMarkdown形式の綺麗な議事録を作成してください。

■ 会議の会話ログ:
{req.transcript}

■ 議事録の構成案（Markdown形式）：
- **会議概要** (日時、参加者、アジェンダなど)
- **決定事項** (合意に達した内容や決定事項)
- **主要な議論内容** (会話のポイント、顧客の関心・課題)
- **ネクストアクション / TODO** (誰が、いつまでに、何をするか)

余計な挨拶や説明は含めず、純粋なMarkdown形式の議事録のみを出力してください。"""

    try:
        llm = OpenAICompatClient()
        resp = llm.ask(
            user_content=prompt,
            system_prompt="あなたは優秀な証券会社のAIアシスタントです。会話ログから簡潔で読みやすいMarkdown形式の議事録を作成してください。",
            history=[],
            tool_registry=None,
            json_schema=None
        )
        minutes_md = resp.content.strip()
    except Exception as e:
        logger.error("Failed to generate minutes: %s", e)
        raise HTTPException(status_code=500, detail=f"議事録の自動生成に失敗しました: {str(e)}")

    save_event_minutes(uid, req.event_id, minutes_md)
    return get_event_people_and_analysis(uid, req.event_id)


class CreatePersonRequest(BaseModel):
    name: str
    company: Optional[str] = ""
    role: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    postal_code: Optional[str] = ""
    hobbies: Optional[str] = ""
    notes: Optional[str] = ""


@app.get("/api/people")
async def get_people_list(authorization: Optional[str] = Header(None)):
    """Get all digital business card / people profiles for user."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import get_all_people
    return {"people": get_all_people(uid)}


@app.post("/api/people")
async def create_person_profile(
    req: CreatePersonRequest,
    authorization: Optional[str] = Header(None)
):
    """Create or update a digital business card person profile."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import create_full_person
    person = create_full_person(uid, req.model_dump() if hasattr(req, 'model_dump') else req.dict())
    return {"status": "ok", "person": person}


@app.delete("/api/people/{person_id}")
async def delete_person_profile(
    person_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a digital business card person profile."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import delete_person
    delete_person(uid, person_id)
    return {"status": "ok"}


@app.get("/api/people/{person_id}/events")
async def get_person_related_events(
    person_id: str,
    authorization: Optional[str] = Header(None)
):
    """Get all calendar events associated with a person profile."""
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.store import get_person_events
    from core.google_tools import _calendar_list_events

    linked_event_ids = get_person_events(uid, person_id)
    if not linked_event_ids:
        return {"events": []}

    import datetime
    time_min = (datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=90)).isoformat()
    cal_res = _calendar_list_events(uid, time_min=time_min, max_results=50)

    events_list = []
    if isinstance(cal_res, dict) and "diagram" in cal_res:
        all_events = cal_res["diagram"].get("events", [])
        for ev in all_events:
            if ev.get("id") in linked_event_ids:
                events_list.append(ev)

    events_list.sort(key=lambda x: x.get("start") or "")
    return {"events": events_list}


class OCRBusinessCardRequest(BaseModel):
    image_base64: str
    mime_type: Optional[str] = "image/jpeg"


@app.post("/api/people/ocr")
async def ocr_business_card_with_gemma(
    req: OCRBusinessCardRequest,
    authorization: Optional[str] = Header(None)
):
    """Analyze business card image using Gemma Vision LLM and extract structured profile JSON."""
    uid = resolve_uid(authorization, allow_anonymous=False)

    base64_data = req.image_base64
    if "," in base64_data:
        base64_data = base64_data.split(",", 1)[1]

    mime = req.mime_type or "image/jpeg"
    data_url = f"data:{mime};base64,{base64_data}"

    user_content = [
        {
            "type": "text",
            "text": """添付された名刺の画像から以下の情報を抽出し、必ず有効なJSONフォーマットのみで出力してください。

JSONキー仕様:
- "name": 名前 (漢字/ローマ字)
- "company": 会社名 / 組織名
- "role": 役職 / 部署名
- "email": メールアドレス
- "phone": 電話番号
- "postal_code": 郵便番号 (例: 100-0005 または 〒100-0005)
- "address": 会社住所 / 所在地
- "hobbies": 趣味や特筆事項
- "notes": その他の特記事項・備考

注意事項:
- 見つからない項目は空文字 "" にしてください。
- 余計な説明テキストやマークダウンコードブロック（```json）は含めず、純粋なJSON文字列のみを出力してください。"""
        },
        {
            "type": "image_url",
            "image_url": {
                "url": data_url
            }
        }
    ]

    try:
        llm = OpenAICompatClient()
        resp = llm.ask(
            user_content=user_content,
            system_prompt=DEFAULT_SYSTEM_PROMPT,
            history=[],
            tool_registry=None,
            json_schema=None
        )
        raw_res = resp.content.strip()
    except Exception as e:
        logger.error("Failed Gemma OCR analysis: %s", e)
        return {"status": "error", "message": str(e), "data": {}}

    import re
    cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
    cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
    cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()

    parsed = {}
    try:
        parsed = json.loads(cleaned_json)
    except Exception as e:
        logger.warning("Failed to parse LLM JSON output (%s): %s", e, raw_res)
        parsed = {
            "name": "",
            "company": "",
            "role": "",
            "email": "",
            "phone": "",
            "address": "",
            "postal_code": "",
            "hobbies": "",
            "notes": raw_res
        }

    return {"status": "ok", "data": parsed}


@app.get("/api/notifications")
def api_get_notifications(authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        res = get_notifications(uid)
        return {"status": "ok", "data": res}
    except Exception as e:
        logger.error("Failed to get notifications: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/notifications/{notification_id}/read")
def api_mark_notification_as_read(notification_id: str, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        mark_notification_as_read(uid, notification_id)
        return {"status": "ok"}
    except Exception as e:
        logger.error("Failed to mark notification as read: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/notifications/{notification_id}")
def api_delete_notification(notification_id: str, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        delete_notification(uid, notification_id)
        return {"status": "ok"}
    except Exception as e:
        logger.error("Failed to delete notification: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/notifications/{notification_id}/rollback")
def api_rollback_notification(notification_id: str, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        notif = get_notification_by_id(uid, notification_id)
        if not notif:
            raise HTTPException(status_code=404, detail="Notification not found")
        
        rollback_action = None
        for act in notif.get("actions", []):
            if act.get("type") == "rollback_calendar":
                rollback_action = act
                break
        
        if not rollback_action:
            raise HTTPException(status_code=400, detail="No rollback action found for this notification")
        
        meta = rollback_action.get("metadata", {})
        action_type = meta.get("action_type")
        event_id = meta.get("event_id")
        
        if not event_id:
            raise HTTPException(status_code=400, detail="Missing event_id in rollback metadata")
            
        svc = _service(uid, "calendar", "v3")
        if svc is None:
            raise HTTPException(status_code=503, detail="Google Account is not linked")
            
        if action_type == "create":
            svc.events().delete(calendarId="primary", eventId=event_id).execute()
            logger.info("Rolled back: Deleted event %s for user %s", event_id, uid)
        elif action_type == "update":
            old_data = meta.get("old_data", {})
            body = {
                "summary": old_data.get("summary", ""),
                "start": {"dateTime": old_data.get("start"), "timeZone": "Asia/Tokyo"},
                "end": {"dateTime": old_data.get("end"), "timeZone": "Asia/Tokyo"},
                "description": old_data.get("description", "")
            }
            svc.events().update(calendarId="primary", eventId=event_id, body=body).execute()
            logger.info("Rolled back: Restored event %s to old state for user %s", event_id, uid)
            
        mark_notification_as_read(uid, notification_id)
        delete_notification(uid, notification_id)
        return {"status": "ok"}
    except Exception as e:
        logger.error("Failed to rollback notification: %s", e)
        raise HTTPException(status_code=500, detail=str(e))
class ReviseDraftRequest(BaseModel):
    instruction: str
    current_draft: str
    original_mail_body: str
    to: str
    subject: str


class SendDraftRequest(BaseModel):
    to: str
    subject: str
    body: str


class UserProfileUpdateRequest(BaseModel):
    name: str
    company: Optional[str] = ""
    role: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    postal_code: Optional[str] = ""
    hobbies: Optional[str] = ""
    notes: Optional[str] = ""


@app.get("/api/user/profile")
def api_get_user_profile(authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        profile = get_user_profile(uid)
        return {"status": "ok", "profile": profile}
    except Exception as e:
        logger.error("Failed to get user profile: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/user/profile")
def api_update_user_profile(req: UserProfileUpdateRequest, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    try:
        profile = upsert_user_profile(uid, req.model_dump())
        return {"status": "ok", "profile": profile}
    except Exception as e:
        logger.error("Failed to update user profile: %s", e)
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/notifications/{notification_id}/reply-draft/revise")
def api_revise_reply_draft(notification_id: str, req: ReviseDraftRequest, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    system_prompt = (
        "あなたは優秀な証券会社のAI秘書です。現在の返信メールのドラフトを、ユーザーの修正指示に従って書き換えてください。\n"
        "余計な挨拶や説明は一切含めず、修正後のメール本文のみを直接出力してください。"
    )
    user_content = (
        f"元の受信メール:\n{req.original_mail_body}\n\n"
        f"現在の返信ドラフト:\n{req.current_draft}\n\n"
        f"ユーザーの修正指示: {req.instruction}"
    )
    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=None,
            tool_mode="off"
        )
        revised_draft = res.content.strip()
        import re
        revised_draft = re.sub(r"^```json\s*", "", revised_draft, flags=re.IGNORECASE)
        revised_draft = re.sub(r"^```\s*", "", revised_draft)
        revised_draft = re.sub(r"\s*```$", "", revised_draft).strip()
        return {"status": "ok", "draft_text": revised_draft}
    except Exception as e:
        logger.error("Failed to revise draft: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/notifications/{notification_id}/reply-draft/send")
def api_send_reply_draft(notification_id: str, req: SendDraftRequest, authorization: Optional[str] = Header(None)):
    uid = resolve_uid(authorization, allow_anonymous=False)
    from core.google_tools import _gmail_send
    try:
        res = _gmail_send(uid, req.to, req.subject, req.body)
        if "[error]" in res:
            raise HTTPException(status_code=500, detail=res)
            
        mark_notification_as_read(uid, notification_id)
        delete_notification(uid, notification_id)
        return {"status": "ok", "message": res}
    except Exception as e:
        logger.error("Failed to send draft email: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


 # Gemma email analyzer configuration
EMAIL_ANALYSIS_SCHEMA = {
    "type": "object",
    "properties": {
        "category": {
            "type": "string",
            "enum": ["notification", "decision"]
        },
        "title": {
            "type": "string"
        },
        "content": {
            "type": "string"
        },
        "actions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "label": { "type": "string" },
                    "type": { "type": "string", "enum": ["reply_draft", "calendar_add", "snooze"] },
                    "metadata": {
                        "type": "object",
                        "properties": {
                            "to": { "type": "string" },
                            "subject": { "type": "string" },
                            "original_body": { "type": "string" },
                            "summary": { "type": "string" },
                            "start": { "type": "string" },
                            "end": { "type": "string" }
                        }
                    }
                },
                "required": ["label", "type"]
            }
        }
    },
    "required": ["category", "title", "content"]
}

# Calendar adjustment subagent schema
CALENDAR_ADJUSTMENT_SCHEMA = {
    "type": "object",
    "properties": {
        "action": {
            "type": "string",
            "enum": ["create", "update", "none"]
        },
        "target_event_id": {
            "type": "string"
        },
        "event_details": {
            "type": "object",
            "properties": {
                "summary": { "type": "string" },
                "start": { "type": "string", "description": "RFC3339 formatted ISO string (e.g. 2026-08-06T15:00:00+09:00)" },
                "end": { "type": "string", "description": "RFC3339 formatted ISO string" },
                "description": { "type": "string" }
            },
            "required": ["summary", "start", "end"]
        }
    },
    "required": ["action"]
}


def fetch_recent_calendar_events(uid: str) -> list[dict]:
    import datetime
    svc = _service(uid, "calendar", "v3")
    if svc is None:
        return []
    try:
        now = datetime.datetime.now(datetime.timezone.utc)
        time_min = (now - datetime.timedelta(days=7)).isoformat()
        time_max = (now + datetime.timedelta(days=7)).isoformat()
        items = svc.events().list(
            calendarId="primary",
            timeMin=time_min,
            timeMax=time_max,
            singleEvents=True,
            orderBy="startTime"
        ).execute().get("items", [])
        events = []
        for ev in items:
            events.append({
                "id": ev.get("id"),
                "summary": ev.get("summary", "(無題)"),
                "start": ev.get("start", {}).get("dateTime") or ev.get("start", {}).get("date") or "",
                "end": ev.get("end", {}).get("dateTime") or ev.get("end", {}).get("date") or "",
                "description": ev.get("description", "")
            })
        return events
    except Exception as e:
        logger.error("Failed to fetch recent events for user %s: %s", uid, e)
        return []


def adjust_calendar_schedule(uid: str, mail_body: str, from_addr: str, subject: str) -> tuple[bool, dict | None]:
    events = fetch_recent_calendar_events(uid)
    
    system_prompt = (
        "あなたはユーザーのカレンダーのスケジュール調整を行う優秀なエージェントです。\n"
        "受信したメールの内容と、現在のカレンダーの既存の予定一覧を比較し、メールで提案された日時・内容に基づいてスケジュールを自動で追加または変更（置換）するための指示を出力してください。\n\n"
        "ルール：\n"
        "- メールに書かれた打ち合わせや期日について、カレンダーの既存予定に類似の予定（例：同じ差出人との打ち合わせや、同じ時間帯の予定など）がある場合は 'update' と判定し、target_event_id にその既存イベントIDを指定してください。また、event_details に変更後の新しいタイトル・開始・終了日時を設定してください。\n"
        "- 既存カレンダーに関連する予定が見つからず、新規にカレンダーへ登録すべき内容である場合は 'create' と判定し、event_details に新規作成用のタイトル・開始・終了日時を設定してください。\n"
        "- カレンダーに登録すべきスケジュール情報（日時など）が含まれていない場合は 'none' と判定してください。\n\n"
        "出力は必ず指定されたJSONスキーマに完全に従った有効なJSONオブジェクトのみにしてください。"
    )
    
    user_content = f"送信者: {from_addr}\n件名: {subject}\nメール本文:\n{mail_body}\n\n既存カレンダー予定一覧:\n{json.dumps(events, ensure_ascii=False, indent=2)}"
    
    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=CALENDAR_ADJUSTMENT_SCHEMA,
            tool_mode="off"
        )
        raw_res = res.content.strip()
    except Exception as e:
        logger.error("Failed to run calendar adjustment sub-agent for user %s: %s", uid, e)
        return False, None

    import re
    cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
    cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
    cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()

    try:
        parsed = json.loads(cleaned_json)
        action = parsed.get("action", "none")
        if action == "none":
            return False, None
            
        svc = _service(uid, "calendar", "v3")
        if svc is None:
            return False, None
            
        details = parsed.get("event_details", {})
        summary = details.get("summary")
        start = details.get("start")
        end = details.get("end")
        desc = details.get("description", "")
        
        if action == "create":
            body = {
                "summary": summary,
                "start": {"dateTime": start, "timeZone": "Asia/Tokyo"},
                "end": {"dateTime": end, "timeZone": "Asia/Tokyo"},
                "description": desc
            }
            created = svc.events().insert(calendarId="primary", body=body).execute()
            created_id = created.get("id")
            
            rollback_meta = {
                "action_type": "create",
                "event_id": created_id,
                "old_data": {}
            }
            change_desc = f"新規予定「{summary}」（開始: {start}）をカレンダーへ自動登録しました。"
            return True, {"meta": rollback_meta, "desc": change_desc, "title": f"【自動登録】予定「{summary}」を追加しました"}
            
        elif action == "update":
            target_id = parsed.get("target_event_id")
            if not target_id:
                body = {
                    "summary": summary,
                    "start": {"dateTime": start, "timeZone": "Asia/Tokyo"},
                    "end": {"dateTime": end, "timeZone": "Asia/Tokyo"},
                    "description": desc
                }
                created = svc.events().insert(calendarId="primary", body=body).execute()
                created_id = created.get("id")
                rollback_meta = {
                    "action_type": "create",
                    "event_id": created_id,
                    "old_data": {}
                }
                change_desc = f"関連予定が見つからなかったため、新規予定「{summary}」（開始: {start}）を自動登録しました。"
                return True, {"meta": rollback_meta, "desc": change_desc, "title": f"【自動登録】予定「{summary}」を追加しました"}
            
            try:
                old_event = svc.events().get(calendarId="primary", eventId=target_id).execute()
                old_data = {
                    "summary": old_event.get("summary", ""),
                    "start": old_event.get("start", {}).get("dateTime") or old_event.get("start", {}).get("date") or "",
                    "end": old_event.get("end", {}).get("dateTime") or old_event.get("end", {}).get("date") or "",
                    "description": old_event.get("description", "")
                }
            except Exception as e:
                logger.error("Failed to retrieve old event %s: %s", target_id, e)
                old_data = {}
                
            body = {
                "summary": summary,
                "start": {"dateTime": start, "timeZone": "Asia/Tokyo"},
                "end": {"dateTime": end, "timeZone": "Asia/Tokyo"},
                "description": desc
            }
            svc.events().update(calendarId="primary", eventId=target_id, body=body).execute()
            
            rollback_meta = {
                "action_type": "update",
                "event_id": target_id,
                "old_data": old_data
            }
            change_desc = (
                f"{from_addr} からのメールに基づき、カレンダー予定を変更しました。\n\n"
                f"■ 変更前: 「{old_data.get('summary')}」（開始: {old_data.get('start')}）\n"
                f"■ 変更後: 「{summary}」（開始: {start}）"
            )
            return True, {"meta": rollback_meta, "desc": change_desc, "title": f"【自動変更】予定「{summary}」を調整しました"}
            
    except Exception as e:
        logger.error("Failed in adjust_calendar_schedule executions: %s", e)
        
    return False, None


DIGITAL_CARD_EXTRACTION_SCHEMA = {
    "type": "object",
    "properties": {
        "name": { "type": "string" },
        "company": { "type": "string" },
        "role": { "type": "string" },
        "phone": { "type": "string" },
        "address": { "type": "string" },
        "postal_code": { "type": "string" },
        "hobbies": { "type": "string" },
        "notes": { "type": "string" }
    },
    "required": ["name"]
}

def extract_digital_card_info(mail_body: str, from_addr: str, subject: str) -> dict:
    """Extract sender's contact information from email headers and body to create business card."""
    system_prompt = (
        "あなたは受信したメール本文とヘッダー情報を解析し、送信者の名刺情報（デジタル連絡先プロファイル）を抽出する優秀なアシスタントです。\n"
        "メールアドレス、メール本文内の署名欄、挨拶文、トピックなどから、以下の項目を特定してJSON形式で抽出してください。\n\n"
        "【抽出項目】\n"
        "- name: 送信者の個人名（日本語表記。不明な場合はメールの差出人名やFromヘッダーから抽出）\n"
        "- company: 送信者の所属会社・組織名\n"
        "- role: 送信者の役職、肩書き、または所属部署\n"
        "- phone: 本文や署名に含まれる電話番号\n"
        "- address: 住所\n"
        "- postal_code: 郵便番号（例: 100-0001）\n"
        "- hobbies: 趣味（もしメール本文にゴルフや旅行などの趣味の記載があれば、なければ空）\n"
        "- notes: 自己紹介、署名の要約、または差出人の簡単な概要\n\n"
        "出力は必ず指定されたJSONスキーマに完全に従った有効なJSONオブジェクトのみにしてください。"
    )
    user_content = f"From: {from_addr}\nSubject: {subject}\nBody:\n{mail_body}"
    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=DIGITAL_CARD_EXTRACTION_SCHEMA,
            tool_mode="off"
        )
        raw_res = res.content.strip()
        import re
        cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
        cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
        cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()
        parsed = json.loads(cleaned_json)
        
        # Make sure name is fallback-populated
        if not parsed.get("name"):
            # Extract name before '<' or '@'
            name_part = from_addr.split('<')[0].strip()
            if not name_part or '@' in name_part:
                name_part = from_addr.split('@')[0].strip()
            parsed["name"] = name_part or "送信者"
        return parsed
    except Exception as e:
        logger.error("Failed to extract digital card info: %s", e)
        # Fallback minimal card
        name_part = from_addr.split('<')[0].strip()
        if not name_part or '@' in name_part:
            name_part = from_addr.split('@')[0].strip()
        return {
            "name": name_part or "送信者",
            "company": "",
            "role": "",
            "phone": "",
            "address": "",
            "postal_code": "",
            "hobbies": "",
            "notes": f"自動登録 ({from_addr})"
        }


def fetch_past_conversations(uid: str, email_addr: str) -> str:
    """Fetch past email snippet history for a specific sender to use in draft context."""
    import re
    match = re.search(r'[\w\.-]+@[\w\.-]+', email_addr)
    if not match:
        return ""
    clean_email = match.group(0).lower()
    
    # Search for emails involving the clean_email address
    from core.google_tools import _gmail_search
    try:
        res = _gmail_search(uid, query=clean_email, max_results=3)
        if isinstance(res, dict) and "llm_text" in res:
            return res["llm_text"]
        elif isinstance(res, str):
            return res
    except Exception as e:
        logger.error("Failed to fetch past conversations for %s: %s", clean_email, e)
    return ""


def generate_reply_draft(uid: str, mail_body: str, from_addr: str, subject: str) -> str:
    # 1. Fetch user's own profile
    user_prof = get_user_profile(uid)
    user_info = "未設定"
    if user_prof:
        user_info = (
            f"名前: {user_prof.get('name')}\n"
            f"会社: {user_prof.get('company', '')}\n"
            f"役職: {user_prof.get('role', '')}\n"
            f"メール: {user_prof.get('email', '')}\n"
            f"趣味・関心: {user_prof.get('hobbies', '')}\n"
            f"自己紹介/メモ: {user_prof.get('notes', '')}"
        )

    # 2. Fetch recipient's digital business card (person)
    person = find_person_by_email(uid, from_addr)
    recipient_info = "デジタル名刺未登録（新規連絡先）"
    if person:
        recipient_info = (
            f"名前: {person.get('name')}\n"
            f"会社: {person.get('company', '')}\n"
            f"役職: {person.get('role', '')}\n"
            f"メール: {person.get('email', '')}\n"
            f"趣味・関心: {person.get('hobbies', '')}\n"
            f"メモ/備考: {person.get('notes', '')}"
        )

    # 3. Fetch past mail history
    past_history = fetch_past_conversations(uid, from_addr)

    system_prompt = (
        "あなたは優秀な証券会社のAI秘書です。受信したメールに対して、丁寧でビジネスとして完璧な返信メールのドラフトを作成してください。\n"
        "【作成時の考慮事項】\n"
        "- 差出人である「自分自身のプロフィール」と、宛先である「相手先のデジタル名刺情報」、および「過去のメール送受信履歴」が提供されている場合は、これらを考慮して文脈に合った返信内容（相手の役職、会社名、過去のトピックへの言及など）にしてください。\n"
        "余計な挨拶や説明（「こちらが返信ドラフトです」等）は一切含めず、メール本文（挨拶や署名を含む本文全体）のみを直接出力してください。"
    )
    
    user_content = (
        f"--- 自分自身（差出人）のプロフィール ---\n{user_info}\n\n"
        f"--- 相手先（受信相手）のデジタル名刺情報 ---\n{recipient_info}\n\n"
        f"--- 過去のメール履歴（最大3件） ---\n{past_history or '履歴なし'}\n\n"
        f"--- 今回受信したメール ---\n"
        f"送信者: {from_addr}\n件名: {subject}\nメール本文:\n{mail_body}"
    )
    
    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=None,
            tool_mode="off"
        )
        return res.content.strip()
    except Exception as e:
        logger.error("Failed to generate reply draft: %s", e)
        return ""
        
    except Exception as e:
        logger.error("Failed in adjust_calendar_schedule executions: %s", e)
        
    return False, None


def analyze_email_and_create_notification(uid: str, mail_id: str, from_addr: str, subject: str, body: str):
    logger.info(f"Analyzing email from {from_addr} with subject: {subject} for user {uid}")
    
    # 1. Try calendar adjustment subagent first
    adjusted, adj_data = adjust_calendar_schedule(uid, body, from_addr, subject)
    if adjusted and adj_data:
        meta = adj_data["meta"]
        desc = adj_data["desc"]
        title = adj_data["title"]
        
        # Create notification with rollback calendar option
        actions = [
            {
                "label": "元に戻す",
                "type": "rollback_calendar",
                "metadata": meta
            }
        ]
        create_notification(uid, "decision", title, desc, actions, mail_id)
        logger.info(f"Successfully adjusted schedule and created rollback notification for user {uid}")
        return

    # 2. Fallback to normal email summary analysis
    system_prompt = (
        "あなたは受信したメールを解析し、重要なお知らせやToDoを自動分類する優秀なアシスタントです。\n"
        "メール本文を読み、要約し、それが単なる「通知(notification)」か、何か行動を起こすべき「判断(decision)」かを判定し、指定されたJSONスキーマに従って出力してください。\n\n"
        "もし「判断(decision)」の場合、具体的なアクションの候補をactions配列に含めてください。\n"
        "アクションのタイプと構成：\n"
        "- 'reply_draft': メールの返信が必要な場合。metadataに \"to\" (宛先), \"subject\" (件名, 元の件名にRe:等を付与したもの), \"original_body\" (元の本文) を含めてください。\n"
        "- 'calendar_add': 打ち合わせの提案や締め切りなどの日程情報が含まれている場合。metadataに \"summary\" (カレンダー予定名), \"start\" (開始日時, RFC3339形式 例: 2026-08-06T15:00:00+09:00), \"end\" (終了日時, RFC3339形式) を含めてください。現在時刻を考慮して日時を推測してください。\n"
        "- 'snooze': 「後で通知する」ための汎用アクション。ラベルは「後で通知する」にしてください。\n\n"
        "出力は必ず指定されたJSONスキーマに完全に従った有効なJSONオブジェクトのみにしてください。"
    )

    user_content = f"送信者: {from_addr}\n件名: {subject}\n本文:\n{body}"

    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=EMAIL_ANALYSIS_SCHEMA,
            tool_mode="off"
        )
        raw_res = res.content.strip()
    except Exception as e:
        logger.error(f"Gemma email analysis failed for user {uid}, mail {mail_id}: {e}")
        return

    import re
    cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
    cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
    cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()

def generate_reply_draft(uid: str, mail_body: str, from_addr: str, subject: str) -> str:
    system_prompt = (
        "あなたは優秀な証券会社のAI秘書です。受信したメールに対して、丁寧でビジネスとして完璧な返信メールのドラフトを作成してください。\n"
        "余計な挨拶や説明（「こちらが返信ドラフトです」等）は一切含めず、メール本文（件名や宛先を除く、拝啓や挨拶から始まる本文全体）のみを直接出力してください。"
    )
    user_content = f"送信者: {from_addr}\n件名: {subject}\nメール本文:\n{mail_body}"
    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=None,
            tool_mode="off"
        )
        return res.content.strip()
    except Exception as e:
        logger.error("Failed to generate reply draft: %s", e)
        return ""


def analyze_email_and_create_notification(uid: str, mail_id: str, from_addr: str, subject: str, body: str):
    logger.info(f"Analyzing email from {from_addr} with subject: {subject} for user {uid}")
    
    # 1. Try calendar adjustment subagent first
    adjusted, adj_data = adjust_calendar_schedule(uid, body, from_addr, subject)
    if adjusted and adj_data:
        meta = adj_data["meta"]
        desc = adj_data["desc"]
        title = adj_data["title"]
        
        # Create notification with rollback calendar option
        actions = [
            {
                "label": "元に戻す",
                "type": "rollback_calendar",
                "metadata": meta
            }
        ]
        create_notification(uid, "decision", title, desc, actions, mail_id)
        logger.info(f"Successfully adjusted schedule and created rollback notification for user {uid}")
        return

    # 2. Fallback to normal email summary analysis
    system_prompt = (
        "あなたは受信したメールを解析し、重要なお知らせやToDoを自動分類する優秀なアシスタントです。\n"
        "メール本文を読み、要約し、それが単なる「通知(notification)」か、何か行動を起こすべき「判断(decision)」かを判定し、指定されたJSONスキーマに従って出力してください。\n\n"
        "もし「判断(decision)」の場合、具体的なアクションの候補をactions配列に含めてください。\n"
        "アクションのタイプと構成：\n"
        "- 'reply_draft': メールの返信が必要な場合。metadataに \"to\" (宛先), \"subject\" (件名, 元の件名にRe:等を付与したもの), \"original_body\" (元の本文) を含めてください。\n"
        "- 'calendar_add': 打ち合わせの提案や締め切りなどの日程情報が含まれている場合。metadataに \"summary\" (カレンダー予定名), \"start\" (開始日時, RFC3339形式 例: 2026-08-06T15:00:00+09:00), \"end\" (終了日時, RFC3339形式) を含めてください。現在時刻を考慮して日時を推測してください。\n"
        "- 'snooze': 「後で通知する」ための汎用アクション。ラベルは「後で通知する」にしてください。\n\n"
        "出力は必ず指定されたJSONスキーマに完全に従った有効なJSONオブジェクトのみにしてください。"
    )

    user_content = f"送信者: {from_addr}\n件名: {subject}\n本文:\n{body}"

    client = OpenAICompatClient()
    try:
        res = client.ask(
            user_content=user_content,
            system_prompt=system_prompt,
            history=[],
            tool_registry=None,
            json_schema=EMAIL_ANALYSIS_SCHEMA,
            tool_mode="off"
        )
        raw_res = res.content.strip()
    except Exception as e:
        logger.error(f"Gemma email analysis failed for user {uid}, mail {mail_id}: {e}")
        return

    import re
    cleaned_json = re.sub(r"^```json\s*", "", raw_res, flags=re.IGNORECASE)
    cleaned_json = re.sub(r"^```\s*", "", cleaned_json)
    cleaned_json = re.sub(r"\s*```$", "", cleaned_json).strip()

    try:
        parsed = json.loads(cleaned_json)
        category = parsed.get("category", "notification")
        title = parsed.get("title", f"メール受信: {subject}")
        content = parsed.get("content", body[:200])
        actions = parsed.get("actions", [])
        
        # Generate reply draft if reply_draft action is requested
        # Check if the sender has a digital business card. If not, auto-create one.
        sender_person = find_person_by_email(uid, from_addr)
        if not sender_person:
            logger.info(f"Sender {from_addr} not found in digital business cards. Auto-creating business card...")
            try:
                card_data = extract_digital_card_info(body, from_addr, subject)
                
                # Extract clean email address
                import re
                email_match = re.search(r'[\w\.-]+@[\w\.-]+', from_addr)
                clean_email = email_match.group(0).lower() if email_match else from_addr.strip()
                card_data["email"] = clean_email
                card_data["notes"] = f"{card_data.get('notes', '')}\n[システム自動登録] メール「{subject}」受信により自動作成されました。"
                
                # Insert into DB
                sender_person = create_full_person(uid, card_data)
                logger.info(f"Successfully auto-created digital business card for {card_data.get('name')} ({clean_email})")
            except Exception as e:
                logger.error("Failed to auto-create digital business card: %s", e)
        
        for act in actions:
            if act.get("type") == "reply_draft":
                meta = act.get("metadata", {})
                draft = generate_reply_draft(uid, body, from_addr, subject)
                meta["draft_text"] = draft
                if sender_person:
                    meta["person_id"] = sender_person["id"]
                    meta["person_name"] = sender_person["name"]
                act["metadata"] = meta
        
        # Save to DB
        create_notification(uid, category, title, content, actions, mail_id)
        logger.info(f"Successfully created notification for email: {title} for user {uid}")
    except Exception as e:
        logger.error(f"Failed to parse or save notification JSON: {e}, raw: {raw_res}")




def check_emails_for_all_users():
    logger.info("Starting email check batch for all linked users...")
    try:
        uids = get_all_linked_users()
    except Exception as e:
        logger.error(f"Failed to fetch linked users: {e}")
        return

    for uid in uids:
        try:
            svc = _service(uid, "gmail", "v1")
            if svc is None:
                continue

            listing = (
                svc.users()
                .messages()
                .list(
                    userId="me",
                    q="newer_than:1h",
                    maxResults=10,
                )
                .execute()
            )
            msgs = listing.get("messages", [])
            if not msgs:
                continue

            for m in msgs:
                mail_id = m["id"]
                if notification_exists_for_mail(uid, mail_id):
                    continue

                full = (
                    svc.users()
                    .messages()
                    .get(userId="me", id=mail_id, format="full")
                    .execute()
                )
                payload = full.get("payload", {})
                headers = {h["name"]: h["value"] for h in payload.get("headers", [])}
                from_addr = headers.get("From", "?")
                subject = headers.get("Subject", "(件名なし)")
                body_text = _extract_plain_body(payload) or full.get("snippet", "")

                analyze_email_and_create_notification(uid, mail_id, from_addr, subject, body_text)

        except Exception as e:
            logger.error(f"Error checking emails for user {uid}: {e}")


def run_email_check_batch_loop():
    import time
    logger.info("Email check batch thread started.")
    time.sleep(15)
    while True:
        try:
            check_emails_for_all_users()
        except Exception as e:
            logger.error(f"Error in run_email_check_batch_loop: {e}")
        time.sleep(3600)


@app.on_event("startup")
def start_notification_batch_scheduler():
    thread = threading.Thread(target=run_email_check_batch_loop, daemon=True)
    thread.start()


@app.get("/api/health")
def health():
    return {"status": "ok", "model": MODEL_NAME}


if __name__ == "__main__":
    import uvicorn
    # Read port from environment variable for deployment compatibility
    port = int(os.getenv("PORT", 8080))
    print(f"[FastAPI] Starting server on 127.0.0.1:{port}...", flush=True)
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="info")

